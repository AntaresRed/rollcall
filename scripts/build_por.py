#!/usr/bin/env python3
"""
Build src/data/por.json — the positions-of-responsibility contact lists.

    python3 scripts/build_por.py "data/POR Contacts Sheet.xlsx" src/data/por.json

Eight sheets in one workbook, in four different shapes:

  Student Council            Post / Name / Phone / Email
  Preparation Committee      Name / Contact            (no post at all)
  Placement Representatives  Name / Contact            (a title row first)
  Clubs, SIGs, Chapters      Group / Post / Name / Email / Contact, where the
                             group cell is filled only on its first row
  Cultural Cell              Post / Name / Email / Contact
  Sports Council             Vertical / Name / Contact, then a "Sports
                             Captains" divider and a *second* header row

Rather than eight parsers, each sheet is described below by where its columns
are and how its rows group, and one reader handles all of them. A new sheet
next year is a table entry, not new code.

Every output row is the same shape — section, role, name, phone, email — so
the screen renders all seven datasets identically.
"""

import json
import re
import sys

import openpyxl

# ---------------------------------------------------------------- corrections
#
# Declarative, and verified on every run: if the workbook is reissued with one
# of these already fixed — or with the cell changed to something else entirely
# — the build fails rather than silently reapplying a stale correction. Same
# reasoning as data/overrides.json for the class catalogue.
#
# Each was found by cross-checking every number against the student roll in
# SplitContacts62.csv. The roll is the more recently verified source, so it
# wins where the two disagree AND the difference looks like a slip rather than
# a deliberately different number. Sequential numbers that clearly belong to an
# issued block — the Placement Representatives' 7605026260-69 — are left alone.
CORRECTIONS = [
    # (sheet, name, wrong number, corrected number, why)
    ("Student Council", "sankeerthana valaboju", "630344607", "6303444607",
     "nine digits in the sheet; a digit was dropped"),
    ("SIGs", "T. Algeria Kom", "7449382453", "8787415552",
     "swapped with Sourav Deb's number"),
    ("SIGs", "Sourav Deb", "8787415552", "7449382453",
     "swapped with T. Algeria Kom's number"),
    ("SIGs", "Yash Vinayak Patil", "9324277117", "9324377117",
     "one digit adrift from the roll"),
]

# ------------------------------------------------------------------ additions
#
# People the workbook lists on one sheet who should also appear on another.
# The Placement Representatives are on the Student Council in their own right,
# but the council's own sheet does not carry them, so without this they show up
# only under CDPO and someone reading the council list would conclude the post
# is vacant.
#
# Name and number are copied from the source sheet rather than repeated here:
# there is one place a rep's number is written down, so the two lists cannot
# drift apart. Like CORRECTIONS, this is verified on every run — a name that
# has left the source sheet, or that the target sheet has since started
# carrying itself, fails the build rather than being silently skipped.
#
# Full names, deliberately: "Abhishek Kumar" is also a SIG head with a
# different number, and picking the wrong one would be invisible.
ADDITIONS = [
    # (source sheet, name, target sheet, role to give them there)
    ("Placement Representatives", "Hetav Hiten Shah", "Student Council", "Placement Representative"),
    ("Placement Representatives", "Nikhil Yadav", "Student Council", "Placement Representative"),
    ("Placement Representatives", "Naveen Yadav", "Student Council", "Placement Representative"),
    ("Placement Representatives", "Thoutam Mahonnath", "Student Council", "Placement Representative"),
    ("Placement Representatives", "Abhishek Kumar", "Student Council", "Placement Representative"),
    ("Placement Representatives", "Amit Prasad", "Student Council", "Placement Representative"),
]

# ---------------------------------------------------------------- sheet shapes
#
# `group` is the column carrying a club/SIG/chapter name, filled only on the
# first row of each group and blank on the rest — so it is carried downwards.
# `skip_until_header` handles the sheets that open with a decorative title row
# before their real header.
SHEETS = {
    "Student Council": dict(role=0, name=1, phone=2, email=3, group=None),
    "Preparation Committee": dict(role=None, name=0, phone=1, email=None, group=None),
    "Placement Representatives": dict(role=None, name=0, phone=1, email=None, group=None),
    "Clubs": dict(role=1, name=2, phone=4, email=3, group=0),
    "SIGs": dict(role=1, name=2, phone=4, email=3, group=0),
    "Chapters": dict(role=1, name=2, phone=4, email=3, group=0),
    "Cultural Cell": dict(role=0, name=1, phone=3, email=2, group=None),
    "Sports Council": dict(role=0, name=1, phone=2, email=None, group=None),
}

# Cells that are a header or a decorative title rather than a person.
HEADER_WORDS = {
    "post", "name", "phone", "email", "email id", "contact", "contact no",
    "contact number", "club", "sig", "chapter", "vertical", "sport", "captain",
}

PHONE_RE = re.compile(r"^[6-9]\d{9}$")


def cell(v):
    """Excel hands a phone back as the float 7252895480.0."""
    s = "" if v is None else str(v).strip()
    return s[:-2] if re.fullmatch(r"\d+\.0", s) else s


def phone_of(raw):
    d = re.sub(r"\D", "", raw or "")
    if len(d) == 12 and d.startswith("91"):
        d = d[2:]
    return d if PHONE_RE.match(d) else None


def is_header_row(cells):
    filled = [c.lower() for c in cells if c]
    return bool(filled) and all(c in HEADER_WORDS for c in filled)


def read_sheet(wb, sheet):
    """One sheet -> [{section, role, name, phone, email}], plus its warnings.

    A row with no name is either a header, a decorative title, or the
    "Sports Captains" divider that starts a new section mid-sheet. All three
    are handled here rather than by a special case per sheet.
    """
    spec = SHEETS[sheet]
    rows, warnings = [], []
    section = None          # the divider a row currently sits under
    group = None            # the carried-down club/SIG/chapter name

    for raw in wb[sheet].iter_rows(values_only=True):
        cells = [cell(c) for c in raw]
        while cells and cells[-1] == "":
            cells.pop()
        if not any(cells):
            continue
        if is_header_row(cells):
            continue

        at = lambda i: cells[i] if i is not None and i < len(cells) else ""

        if spec["group"] is not None and at(spec["group"]):
            group = at(spec["group"])

        name = at(spec["name"])
        phone_raw = at(spec["phone"])

        # A row holding one lone cell and no number is a caption, not a person.
        # Which kind of caption depends on where it sits: before any person it
        # is the sheet's own title ("Placement Representatives", "Cultural
        # cell"), which would otherwise be read as somebody called that,
        # because on those sheets the title lands in the name column. After
        # people have been read it is a divider — "Sports Captains" — and
        # starts a new section.
        only = [c for c in cells if c]
        if len(only) == 1 and not phone_of(phone_raw):
            if rows:
                section = only[0]
            continue
        if not name:
            continue

        phone = phone_of(phone_raw)
        email = at(spec["email"]) or None
        if email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[a-z]{2,}", email, re.I):
            warnings.append(f"{sheet}: {name} — email looks malformed: {email!r}")

        rows.append({
            "section": group or section,
            "role": at(spec["role"]) or None,
            "name": name,
            "phone": phone,
            "email": email,
            # Kept only until corrections have run. A number broken badly
            # enough to fail normalisation — nine digits, say — arrives here as
            # None, so a correction targeting it has nothing to match on
            # unless the original cell is still around.
            "_raw": re.sub(r"\D", "", phone_raw or ""),
        })

    return rows, warnings


def apply_corrections(by_sheet):
    """Rewrite the four known-bad numbers, and insist each one still applies."""
    applied, stale = [], []
    for sheet, name, wrong, right, why in CORRECTIONS:
        hit = next(
            (r for r in by_sheet.get(sheet, [])
             if r["name"].strip().lower() == name.strip().lower()
             and wrong in ((r["phone"] or ""), r.get("_raw", ""))),
            None,
        )
        if hit is None:
            stale.append((sheet, name, wrong, right, why))
            continue
        hit["phone"] = right
        applied.append((sheet, name, wrong, right, why))
    return applied, stale


def apply_additions(by_sheet):
    """Copy the cross-listed people onto their second sheet.

    Runs after the corrections, so a number repaired above is the one that gets
    copied — not the raw cell it was repaired from.
    """
    applied, problems = [], []
    for source, name, target, role in ADDITIONS:
        matches = [r for r in by_sheet.get(source, [])
                   if r["name"].strip().lower() == name.strip().lower()]
        if len(matches) != 1:
            problems.append(
                f"{name!r}: expected exactly one on the {source} sheet, found {len(matches)}")
            continue
        if any(r["name"].strip().lower() == name.strip().lower()
               and (r["role"] or "").strip().lower() == role.strip().lower()
               for r in by_sheet.get(target, [])):
            problems.append(
                f"{name!r}: the {target} sheet now lists them as {role} itself — "
                f"drop this entry from ADDITIONS")
            continue
        src = matches[0]
        by_sheet[target].append({
            "section": None,
            "role": role,
            "name": src["name"],
            "phone": src["phone"],
            "email": src["email"],
        })
        applied.append((source, name, target, role))
    return applied, problems


def sectioned(rows, kind=None, first_label=None):
    """Group flat rows into the sections the screen renders, order preserved.

    `first_label` names the opening run of rows on a sheet that only starts
    using dividers partway down — Sports Council lists the council itself with
    no heading, then says "Sports Captains" and carries on.
    """
    out = []
    for r in rows:
        label = r["section"]
        if not out or out[-1]["label"] != label:
            out.append({"label": label, "kind": kind, "people": []})
        out[-1]["people"].append(
            {k: r[k] for k in ("role", "name", "phone", "email")})
    if out and out[0]["label"] is None and first_label and len(out) > 1:
        out[0]["label"] = first_label
    return out


def build(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"workbook is missing sheet(s): {', '.join(missing)}")

    by_sheet, warnings = {}, []
    for sheet in SHEETS:
        rows, warn = read_sheet(wb, sheet)
        by_sheet[sheet] = rows
        warnings += warn

    applied, stale = apply_corrections(by_sheet)

    added, add_problems = apply_additions(by_sheet)
    if add_problems:
        raise SystemExit(
            "ADDITIONS no longer describe the workbook:\n  "
            + "\n  ".join(add_problems))

    # Reported only now: a number the corrections above have already repaired
    # is not something anyone needs to go and look at.
    for sheet, rows in by_sheet.items():
        for r in rows:
            if r["phone"] is None and r.get("_raw"):
                warnings.append(
                    f"{sheet}: {r['name']} — unusable number {r['_raw']!r}, left blank")

    # The seven datasets the screens ask for. "SIGs and Chapters" is the only
    # one that merges two sheets; each half keeps a tag so a SIG is still
    # distinguishable from a Chapter once they sit in one list.
    datasets = {
        "student-council": {
            "label": "Student Council",
            "sections": sectioned(by_sheet["Student Council"]),
        },
        "preparation-committee": {
            "label": "Preparation Committee",
            "sections": sectioned(by_sheet["Preparation Committee"]),
        },
        "placement-representatives": {
            "label": "Placement Representatives",
            "sections": sectioned(by_sheet["Placement Representatives"]),
        },
        "clubs": {
            "label": "Clubs",
            "sections": sectioned(by_sheet["Clubs"]),
        },
        "sigs-chapters": {
            "label": "SIGs and Chapters",
            "sections": sectioned(by_sheet["SIGs"], kind="SIG")
                      + sectioned(by_sheet["Chapters"], kind="Chapter"),
        },
        "cultural-cell": {
            "label": "Cultural Cell",
            "sections": sectioned(by_sheet["Cultural Cell"]),
        },
        "sports-council": {
            "label": "Sports Council and Captains",
            "sections": sectioned(by_sheet["Sports Council"], first_label="Sports Council"),
        },
    }
    return datasets, warnings, applied, stale, added


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "data/POR Contacts Sheet.xlsx"
    dest = sys.argv[2] if len(sys.argv) > 2 else "src/data/por.json"

    datasets, warnings, applied, stale, added = build(src)

    with open(dest, "w", encoding="utf-8", newline="\n") as fh:
        json.dump(datasets, fh, indent=2, ensure_ascii=False)

    total = sum(len(p["people"]) for d in datasets.values() for p in d["sections"])
    print(f"{len(datasets)} lists, {total} people -> {dest}")

    if added:
        print(f"\n{len(added)} cross-listed onto a second sheet:")
        for source, name, target, role in added:
            print(f"   {name} — {source} -> {target} as {role}")
    for key, d in datasets.items():
        n = sum(len(s["people"]) for s in d["sections"])
        print(f"   {d['label']:<28} {n:>3} people in {len(d['sections'])} section(s)")

    if applied:
        print(f"\n{len(applied)} correction(s) applied:")
        for sheet, name, wrong, right, why in applied:
            print(f"   {sheet}: {name}  {wrong} -> {right}  ({why})")

    if stale:
        raise SystemExit(
            f"\n{len(stale)} correction(s) no longer match the workbook. Either the "
            f"sheet was fixed at source — in which case delete the entry from "
            f"CORRECTIONS — or the cell changed and the correction needs rechecking:\n"
            + "\n".join(f"   {s}: {n} (expected {w!r})" for s, n, w, _, _ in stale)
        )

    if warnings:
        print(f"\n{len(warnings)} thing(s) worth a look:")
        for w in warnings:
            print(f"   {w}")
