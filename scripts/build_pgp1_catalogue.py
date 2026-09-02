#!/usr/bin/env python3
"""
Build the PGP1 (first year) catalogue.

    python3 scripts/build_pgp1_catalogue.py \
        "data/PGP1 term 2.xlsx" src/data/catalogue-pgp1.json

A sibling of build_catalogue.py rather than a branch inside it. The two
workbooks share almost nothing:

                        PGP2 (Term V)              PGP1 (Term II)
  schedule sheets       one, whole term            two — phase IS the sheet
  day column            A                          B
  time slots            six, incl. 18:00           five, no 18:00
  a row means           parallel electives         nothing; see below
  section               per course, optional       inside the cell text
  instructors           a separate .tsv            an "Instructors" sheet

and one parser trying to be both is how a change for one year silently breaks
the other.

The important structural difference is the last two. PGP2 is elective-driven:
a student picks courses. PGP1 is a core curriculum where every cell is tagged
with a section — "Operations Research (A)" — and a student's whole timetable
follows from which section they are in. So this emits `kind: "sections"`, and
the app shows a section picker instead of a course search.

Output is otherwise the exact shape build_catalogue.py emits, so the admin
upload, the validator and publish_catalogue all work on it unchanged. Two
fields are additive and optional, absent from PGP2 catalogues:

  * every meeting carries a `room`, because a PGP1 room follows the section
    (A=L-21 … F=N-32), not the course
  * instructor entries carry `sections`, so a Section C student is shown the
    professor who actually teaches them

Everything is checked. A sheet that has moved, a course whose sections do not
cover A–F exactly once, a name that cannot be matched to a course code — each
stops the build rather than producing a half-parsed term. The layout is
expected to be stable, but the institute has reissued sheets before.
"""

import collections
import difflib
import json
import re
import sys
from datetime import date

import openpyxl

# ---------------------------------------------------------------- shared rules
# Identical to build_catalogue.py on purpose: the two years sit the same exams
# under the same attendance rules, and a first year's budget should not differ
# from a second year's because two scripts drifted apart.
SLOT_END = {
    "08:30": "10:00", "10:15": "11:45", "12:00": "13:30",
    "14:30": "16:00", "16:15": "17:45", "18:00": "19:30",
}
CREDIT_RULES = {
    3.0: {"total_classes": 20, "min_pct": 75},
    1.5: {"total_classes": 10, "min_pct": 80},
}

PHASE_SHEETS = {"Pre Mid Term-II": "pre_mid", "Post Mid Term-II": "post_mid"}
INSTRUCTOR_SHEET = "Instructors"
CALENDAR_SHEET = "Calendar"

DAYS = {"MONDAY": 1, "TUESDAY": 2, "WEDNESDAY": 3, "THURSDAY": 4,
        "FRIDAY": 5, "SATURDAY": 6, "SUNDAY": 7}

SECTIONS = "ABCDEF"

# A grid cell: "Organizational Behaviour-II (A) \n(VJ) L-21"
CELL_SECTION = re.compile(r"\(([A-F])\)")
ROOM = re.compile(r"\b([LNM]-\d+|Amphi\s*\([^)]*\))")

# The instructor column, which is free text and has used every one of these:
#   "(Secs-A,C)"  "(Sec-B)"  "(A, B and C)"  "Secs AB"  "Secs.CDEF"
SEC_MARKED = re.compile(r"(?:Secs?\.?-?|Sections?)\s*([A-Za-z][A-Za-z,&.\s]*)", re.I)
SEC_BARE = re.compile(r"\(\s*([A-F](?:\s*(?:,|&|and)\s*[A-F])*)\s*\)", re.I)
# "Session 1-5:", "(1-5, All six secs)" — mapped by session range, not section.
SESSION_RANGED = re.compile(r"Session|\b\d+\s*-\s*\d+\b")

MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], start=1)}


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def key(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def die(msg):
    raise SystemExit(f"PGP1 catalogue: {msg}")


# ---------------------------------------------------------------- the grid

def find_header(ws):
    """Locate the header row and which column each time slot sits in.

    Read rather than hardcoded: the sheet gained a leading blank column
    somewhere between the PGP2 and PGP1 workbooks, and would do it again.
    """
    for r in range(1, 12):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(isinstance(v, str) and "timeslot" in v.lower() for v in row if v):
            slots, day_col = {}, None
            for i, v in enumerate(row, start=1):
                if isinstance(v, str) and "timeslot" in v.lower():
                    day_col = i
                elif isinstance(v, str) and re.fullmatch(r"\d{2}:\d{2}", v.strip()):
                    slots[i] = v.strip()
            if day_col and slots:
                return r, day_col, slots
    die("could not find the 'Dayslot / Timeslot' header row")


def read_grid(ws, phase):
    """Every tagged cell on one phase sheet, as flat meetings."""
    header, day_col, slots = find_header(ws)

    unknown = [s for s in slots.values() if s not in SLOT_END]
    if unknown:
        die(f"{ws.title}: unknown time slot(s) {unknown} — SLOT_END needs them")

    out, day = [], None
    for r in range(header + 1, ws.max_row + 1):
        label = norm(ws.cell(row=r, column=day_col).value).upper()
        if label in DAYS:
            day = DAYS[label]
        for col, start in slots.items():
            raw = ws.cell(row=r, column=col).value
            if not isinstance(raw, str) or not raw.strip():
                continue
            text = norm(raw)
            m = CELL_SECTION.search(text)
            if not m:
                continue          # a note, a title, an abbreviation table
            if day is None:
                die(f"{ws.title} row {r}: a class before any weekday heading")
            room = ROOM.search(text[m.end():])
            out.append({
                "name": norm(text[:m.start()]),
                "section": m.group(1),
                "day": day,
                "start": start,
                "end": SLOT_END[start],
                "room": norm(room.group(1)) if room else None,
                "phase": phase,
                "_where": f"{ws.title} r{r}",
            })
    if not out:
        die(f"{ws.title}: no tagged cells found — has the layout changed?")
    return out


# ---------------------------------------------------------------- instructors

def expand(token):
    """'AB' -> A,B.  'and' -> nothing.  'All' -> nothing.

    Without this, findall over "A, B and C" harvests an A and a D out of the
    word 'and', and "All six secs" yields an A.
    """
    t = token.upper()
    return list(t) if t and all(c in SECTIONS for c in t) else []


def sections_in(chunk):
    m = SEC_MARKED.search(chunk) or SEC_BARE.search(chunk)
    if not m:
        return []
    out = []
    for token in re.split(r"[^A-Za-z]+", m.group(1)):
        out += expand(token)
    return sorted(dict.fromkeys(out))


def read_instructors(ws):
    """code -> {name, credits, teachers:[{name, role, sections}], ranged}"""
    header = None
    for r in range(1, 8):
        row = [norm(ws.cell(row=r, column=c).value).lower() for c in range(1, 6)]
        if any("course code" in v for v in row):
            header = r
            break
    if header is None:
        die(f"{ws.title}: no 'Course Code' header")

    courses = {}
    for r in range(header + 1, ws.max_row + 1):
        code = norm(ws.cell(row=r, column=2).value)
        if not code:
            continue
        name = norm(re.sub(r"\((?:pre|post)[\s-]*mid term\)", "",
                           norm(ws.cell(row=r, column=3).value), flags=re.I))
        credits = ws.cell(row=r, column=4).value
        raw = str(ws.cell(row=r, column=5).value or "")

        # Entries are separated by newlines inside the cell, occasionally " / ".
        ranged = bool(SESSION_RANGED.search(raw))
        teachers, seen = [], set()
        for part in raw.split("\n"):
            for chunk in re.split(r"\s+/\s+", part):
                chunk = norm(chunk)
                if not chunk:
                    continue
                who = re.search(r"Prof\.?\s*([^(/,]+?)\s*(?:\(|,|$)", chunk)
                if not who:
                    continue
                person = norm(who.group(1))
                if person.lower() in seen:
                    continue      # OM-103 names its coordinator twice
                seen.add(person.lower())
                teachers.append({
                    "name": person,
                    "role": "CC" if re.search(r"\(\s*CC\s*\)", chunk, re.I) else None,
                    # A course mapped by session range has no section split to
                    # record; every one of its names applies to every section.
                    "sections": [] if ranged else sections_in(chunk),
                })

        try:
            credits = float(credits)
        except (TypeError, ValueError):
            die(f"{code}: credits {credits!r} is not a number")
        if credits not in CREDIT_RULES:
            die(f"{code}: {credits} credits has no attendance rule "
                f"(known: {sorted(CREDIT_RULES)})")

        courses[code] = {"code": code, "name": name, "credits": credits,
                         "teachers": teachers, "ranged": ranged}
    if not courses:
        die(f"{ws.title}: no courses read")
    return courses


def match_code(grid_name, courses, report):
    """Grid names and the Instructors sheet disagree — 'Managerial Decision
    Modelling' vs 'Decisions', and 'Quantitative' vs the sheet's 'Quantitive'.
    Exact first, then a close match, and never a guess."""
    for code, c in courses.items():
        if key(c["name"]) == key(grid_name):
            return code
    best, score = None, 0.0
    for code, c in courses.items():
        s = difflib.SequenceMatcher(None, key(grid_name), key(c["name"])).ratio()
        if s > score:
            best, score = code, s
    if score >= 0.90:
        line = (f"   {grid_name!r} ~ {courses[best]['name']!r} "
                f"-> {best} (similarity {score:.3f})")
        if line not in report:
            report.append(line)
        return best
    die(f"no course code for {grid_name!r} "
        f"(closest: {courses[best]['name']!r} at {score:.2f}). "
        "The Instructors sheet and the grid may have diverged.")


# ---------------------------------------------------------------- calendar

def parse_range(text):
    """'September 07 to October 09, 2026' -> (date, date). Also '-' ranges."""
    t = norm(text)
    year = re.search(r"(20\d{2})", t)
    if not year:
        die(f"no year in calendar entry {t!r}")
    y = int(year.group(1))
    parts = re.findall(r"([A-Za-z]+)\s+(\d{1,2})", t)
    good = [(m, d) for m, d in parts if m.lower() in MONTHS]
    if len(good) < 2:
        die(f"could not read a date range from {t!r}")
    (m1, d1), (m2, d2) = good[0], good[1]
    return (date(y, MONTHS[m1.lower()], int(d1)),
            date(y, MONTHS[m2.lower()], int(d2)))


def read_calendar(ws):
    rows = {}
    for r in ws.iter_rows(values_only=True):
        cells = [norm(c) for c in r if c is not None and norm(c)]
        if len(cells) >= 2:
            rows[cells[0].lower()] = cells[1]

    def find(*needles):
        for label, value in rows.items():
            if all(n in label for n in needles):
                return value
        die(f"calendar has no row matching {needles}")

    pre = parse_range(find("pre mid"))
    post = parse_range(find("post mid"))

    breaks = []
    for needles, label in [(("summer", "placement"), "Summer placement"),
                           (("puja",), "Puja vacation"),
                           (("mid term exam",), "Mid-term exams"),
                           (("end term exam",), "End-term exams")]:
        raw = find(*needles)
        a, b = parse_range(raw)
        breaks.append({"label": label, "from": a.isoformat(),
                       "to": b.isoformat(), "note": raw})
    breaks.sort(key=lambda b: b["from"])

    return {
        "term_start": pre[0].isoformat(),
        "pre_mid_end": pre[1].isoformat(),
        "post_mid_start": post[0].isoformat(),
        # The last teaching day, not the last day of term — the exam week that
        # follows is a break, exactly as the PGP2 calendar is built.
        "term_end": post[1].isoformat(),
        "breaks": breaks,
    }


# ---------------------------------------------------------------- assembly

def attach_emails(courses, directory_path):
    """Institute addresses, matched by name against the faculty directory."""
    try:
        with open(directory_path, encoding="utf-8") as fh:
            people = json.load(fh)
    except FileNotFoundError:
        return 0, []
    by_name = {}
    for p in people:
        email = next((o.get("email") for o in p.get("offices") or [] if o.get("email")), None)
        if email:
            by_name[key(p["name"])] = email

    hit, missed = 0, []
    for c in courses:
        for t in c["instructors"]:
            email = by_name.get(key(t["name"]))
            if email is None:
                best, score = None, 0.0
                for k, v in by_name.items():
                    s = difflib.SequenceMatcher(None, key(t["name"]), k).ratio()
                    if s > score:
                        best, score = v, s
                email = best if score >= 0.92 else None
            t["email"] = email
            if email:
                hit += 1
            else:
                missed.append(f"{c['code']}: {t['name']}")
    return hit, missed


def build(path, directory_path, cohort_year):
    wb = openpyxl.load_workbook(path, data_only=True)

    missing = [s for s in list(PHASE_SHEETS) + [INSTRUCTOR_SHEET, CALENDAR_SHEET]
               if s not in wb.sheetnames]
    if missing:
        die(f"workbook is missing sheet(s): {', '.join(missing)}")

    meetings = []
    for sheet, phase in PHASE_SHEETS.items():
        meetings += read_grid(wb[sheet], phase)

    catalogue_courses = read_instructors(wb[INSTRUCTOR_SHEET])
    calendar = read_calendar(wb[CALENDAR_SHEET])

    fuzzy = []
    by_code = {}
    for m in meetings:
        code = match_code(m["name"], catalogue_courses, fuzzy)
        by_code.setdefault(code, []).append(m)

    courses = []
    for code, ms in sorted(by_code.items()):
        meta = catalogue_courses[code]
        phases = {m["phase"] for m in ms}
        # A course on both sheets runs the whole term; one sheet, half of it.
        phase = "full" if len(phases) > 1 else next(iter(phases))

        # A slot that appears on BOTH sheets runs all term and is stored once,
        # unphased. A slot on only one sheet belongs to that half and says so.
        #
        # This is not a tidying detail. Operations Research keeps its Monday
        # slot across the term but swaps its second one between the halves, so
        # storing all four as full-term would put a phantom class on the grid
        # every week of the year — and fire an alert for it.
        seen = {}
        for m in ms:
            slot = (m["section"], m["day"], m["start"])
            seen.setdefault(slot, {"m": m, "phases": set()})["phases"].add(m["phase"])

        sections = {}
        for (letter, _, _), entry in seen.items():
            m, phases = entry["m"], entry["phases"]
            meeting = {"day": m["day"], "start": m["start"], "end": m["end"]}
            if m["room"]:
                meeting["room"] = m["room"]
            # Absent means "every teaching week", which is what both halves
            # together amounts to.
            if len(phases) == 1 and len(set(x["phase"] for x in ms)) > 1:
                meeting["phase"] = next(iter(phases))
            sections.setdefault(letter, []).append(meeting)
        for letter in sections:
            sections[letter].sort(key=lambda x: (x["day"], x["start"]))

        held = sorted(sections)
        if held != list(SECTIONS):
            die(f"{code} ({meta['name']}) has sections {''.join(held)}, "
                f"expected {SECTIONS} — the grid may be incomplete")

        # Per-section instructors must partition A–F. If they do not, the
        # mapping was misread and a student would be shown the wrong name.
        if not meta["ranged"]:
            covered = sorted(s for t in meta["teachers"] for s in t["sections"])
            if covered != list(SECTIONS):
                die(f"{code} ({meta['name']}): instructors cover "
                    f"{''.join(covered) or 'nothing'}, expected each of "
                    f"{SECTIONS} exactly once")

        # The grid's spelling, not the Instructors sheet's. The two disagree
        # ("Quantitative" vs "Quantitive"), and the grid is what the institute
        # prints on the timetable and what a student will recognise.
        grid_name = collections.Counter(m["name"] for m in ms).most_common(1)[0][0]

        courses.append({
            "code": code,
            "name": grid_name,
            "phase": phase,
            "credits": meta["credits"],
            **CREDIT_RULES[meta["credits"]],
            # The course-level venue is only meaningful when every section
            # shares one, which for PGP1 it never is — the room follows the
            # section. Kept empty so nothing reads it by accident; the real
            # room is on each meeting.
            "venue": "",
            "schedule_type": "weekly",
            "sections": sections,
            "instructors": [
                {"name": t["name"], "role": t["role"], "sections": t["sections"]}
                for t in meta["teachers"]
            ],
        })

    # Every section sits the same core curriculum, so each must end up with the
    # same number of weekly meetings in each half. Sections legitimately differ
    # in how many slots they keep across the mid-term — that is why the stored
    # counts vary — but a half where one section has fewer classes than another
    # means a cell was misread, and a student would simply never be asked about
    # the class that went missing.
    #
    # Derived rather than hardcoded, so next term's shape does not need a code
    # change to be checked.
    weekly = {}
    for phase in ("pre_mid", "post_mid"):
        weekly[phase] = {}
        for letter in SECTIONS:
            n = 0
            for c in courses:
                for m in c["sections"][letter]:
                    if (m.get("phase") or c["phase"]) in (phase, "full"):
                        n += 1
            weekly[phase][letter] = n
        counts = set(weekly[phase].values())
        if len(counts) > 1:
            detail = "  ".join(f"{k}={v}" for k, v in weekly[phase].items())
            die(f"sections disagree on how many {phase} classes a week they "
                f"have ({detail}). One of them is missing a class — recheck "
                "the grid before publishing this.")

    unused = sorted(set(catalogue_courses) - set(by_code))

    print(f"weekly per section: {weekly['pre_mid']['A']} pre-mid, "
          f"{weekly['post_mid']['A']} post-mid — identical across {SECTIONS}")

    return {
        "term": f"TERM-II: AY {calendar['term_start'][:4]}-"
                f"{int(calendar['term_start'][:4]) + 1}",
        "cohort_year": cohort_year,
        "kind": "sections",
        "slots": sorted({m["start"] for m in meetings}),
        "calendar": calendar,
        "courses": courses,
    }, fuzzy, unused


# ---------------------------------------------------------------- main

if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "data/PGP1 term 2.xlsx"
    dest = sys.argv[2] if len(sys.argv) > 2 else "src/data/catalogue-pgp1.json"
    directory = sys.argv[3] if len(sys.argv) > 3 else "src/data/directory.json"
    cohort = int(sys.argv[4]) if len(sys.argv) > 4 else 2028

    catalogue, fuzzy, unused = build(src, directory, cohort)
    hit, missed = attach_emails(catalogue["courses"], directory)

    with open(dest, "w", encoding="utf-8", newline="\n") as fh:
        json.dump(catalogue, fh, indent=2, ensure_ascii=False)

    meetings = sum(len(v) for c in catalogue["courses"] for v in c["sections"].values())
    print(f"{len(catalogue['courses'])} courses, {meetings} meetings "
          f"across sections {SECTIONS} -> {dest}")
    print(f"cohort {catalogue['cohort_year']} · {catalogue['term']} · kind={catalogue['kind']}")

    cal = catalogue["calendar"]
    print(f"\nterm {cal['term_start']} .. {cal['term_end']}  "
          f"(pre-mid ends {cal['pre_mid_end']}, post-mid starts {cal['post_mid_start']})")
    for b in cal["breaks"]:
        print(f"  no class  {b['from']} .. {b['to']}  {b['label']}")

    print("\nper section:")
    for letter in SECTIONS:
        n = sum(len(c["sections"][letter]) for c in catalogue["courses"])
        rooms = sorted({m.get("room") for c in catalogue["courses"]
                        for m in c["sections"][letter] if m.get("room")})
        pre = sum(1 for c in catalogue["courses"] if c["phase"] in ("pre_mid", "full"))
        post = sum(1 for c in catalogue["courses"] if c["phase"] in ("post_mid", "full"))
        print(f"  {letter}: {n} weekly meetings, rooms {', '.join(rooms) or '—'}"
              f"  ({pre} courses pre-mid, {post} post-mid)")

    if fuzzy:
        print("\ncourse names matched by similarity, not exactly — check these:")
        for f in fuzzy:
            print(f)

    print(f"\nfaculty emails: {hit} matched")
    if missed:
        print(f"  no directory entry for {len(missed)}:")
        for m in missed:
            print(f"    {m}")
    if unused:
        print(f"\nin the Instructors sheet but not on the grid: {', '.join(unused)}")
