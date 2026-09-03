#!/usr/bin/env python3
"""
Build src/data/night-menu.json — the night canteen menus.

    python3 scripts/build_night_menu.py \
        "data/Night Mess Menu.xlsx" src/data/night-menu.json

A sibling of build_menu.py, not a branch inside it, because the two menus are
different things wearing the same word. The day mess is a week: seven days,
four meals, no prices. The night canteens are shops: a few hundred priced
items grouped by category, the same list every night.

Sheets:
  Info              Hostel | Canteen | Phone | Hours | Room service (Rs)
  <TAG> Night Menu  Category | Item | Price | Diet

`Diet` is the column the app's veg filter reads. It is data rather than
something derived at render time on purpose: guessing from a name is right
most of the time, and the times it is wrong are the times a vegetarian eats
meat. So the spreadsheet decides, a human can correct it, and anything left
unresolved is carried through as unknown rather than quietly called veg.
"""

import glob
import json
import os
import re
import sys

import openpyxl

DIETS = {"veg", "egg", "non-veg"}
COLUMNS = ["Category", "Item", "Price", "Diet"]


def norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def die(msg):
    raise SystemExit(f"night menu: {msg}")


def read_info(ws):
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(r)]
    if not rows:
        die("Info sheet is empty")
    head = [c.lower() for c in rows[0]]

    def col(*names):
        for i, h in enumerate(head):
            if any(n in h for n in names):
                return i
        return None

    ci = {k: col(*n) for k, n in {
        "id": ("hostel",), "canteen": ("canteen",), "phone": ("phone",),
        "hours": ("hours",), "service": ("room service", "service"),
    }.items()}
    if ci["id"] is None:
        die("Info sheet has no Hostel column")

    out = {}
    for r in rows[1:]:
        tag = r[ci["id"]].upper()
        if not tag:
            continue
        get = lambda k: (r[ci[k]] if ci[k] is not None and ci[k] < len(r) else "")
        out[tag] = {
            "canteen": get("canteen"),
            "phone": get("phone"),
            "hours": get("hours"),
            "room_service": get("service"),
        }
    return out


def read_menu(ws):
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(r)]
    if not rows:
        die(f"{ws.title}: empty")

    head = [c.lower() for c in rows[0]]
    idx = {}
    for want in COLUMNS:
        for i, h in enumerate(head):
            if h == want.lower():
                idx[want] = i
                break
        else:
            die(f"{ws.title}: no '{want}' column (header reads {rows[0]})")

    cats, order, unknown = {}, [], []
    for r in rows[1:]:
        get = lambda k: r[idx[k]] if idx[k] < len(r) else ""
        cat, item, price, diet = (get(c) for c in COLUMNS)
        if not item:
            continue
        if not cat:
            die(f"{ws.title}: '{item}' has no category")
        diet = diet.lower()
        if diet not in DIETS:
            # Carried through as unknown, never silently veg. The screen shows
            # it under every filter and marks it, so nobody is told a dish is
            # vegetarian on the strength of a blank cell.
            unknown.append(f"{cat} / {item}" + (f" ({diet})" if diet else ""))
            diet = "unknown"
        # A clean number stays a number; "22/25" and "7/8" are real prices
        # the canteens print for two sizes, so those stay as written.
        if re.fullmatch(r"\d+(\.\d+)?", price):
            price = int(float(price))

        if cat not in cats:
            cats[cat] = []
            order.append(cat)
        cats[cat].append({"name": item, "price": price, "diet": diet})

    return [{"name": c, "items": cats[c]} for c in order], unknown


def scans(pages_dir, tag):
    """The scanned pages for one canteen, in order.

    Discovered from the directory rather than listed in the spreadsheet:
    adding a page is then dropping `wh-5.jpg` next to the others, which is a
    thing somebody can do without being told a convention twice.
    """
    found = glob.glob(os.path.join(pages_dir, f"{tag.lower()}-*.jpg"))

    def page_no(p):
        m = re.search(r"-(\d+)\.jpg$", os.path.basename(p))
        return int(m.group(1)) if m else 0

    # Served from /menu/night/, which is public/ and so is copied verbatim
    # into the build rather than bundled into the JavaScript. The pages are
    # only fetched when somebody actually asks to see one.
    return [f"/menu/night/{os.path.basename(p)}" for p in sorted(found, key=page_no)]


def build(path, pages_dir="public/menu/night"):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Info" not in wb.sheetnames:
        die("workbook has no Info sheet")
    info = read_info(wb["Info"])

    hostels, warnings = [], []
    for name in wb.sheetnames:
        if name == "Info":
            continue
        tag = norm(name).split()[0].upper()
        cats, unknown = read_menu(wb[name])
        meta = info.get(tag)
        if meta is None:
            die(f"'{name}' has no matching row in the Info sheet (looked for {tag})")
        pages = scans(pages_dir, tag)
        if not pages:
            warnings.append(f"{tag}: no scanned pages in {pages_dir} — "
                            f"the 'see the original' button will not appear")
        hostels.append({"id": tag.lower(), "name": tag, **meta,
                        "pages": pages, "categories": cats})
        warnings += [f"{tag}: {u}" for u in unknown]

    if not hostels:
        die("no menu sheets found")
    return {"hostels": hostels}, warnings


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "data/Night Mess Menu.xlsx"
    dest = sys.argv[2] if len(sys.argv) > 2 else "src/data/night-menu.json"
    pages_dir = sys.argv[3] if len(sys.argv) > 3 else "public/menu/night"

    menu, warnings = build(src, pages_dir)
    with open(dest, "w", encoding="utf-8", newline="\n") as fh:
        json.dump(menu, fh, indent=2, ensure_ascii=False)

    total = sum(len(c["items"]) for h in menu["hostels"] for c in h["categories"])
    print(f"{len(menu['hostels'])} canteens, {total} items -> {dest}")
    for h in menu["hostels"]:
        n = sum(len(c["items"]) for c in h["categories"])
        by = {}
        for c in h["categories"]:
            for i in c["items"]:
                by[i["diet"]] = by.get(i["diet"], 0) + 1
        split = "  ".join(f"{k} {v}" for k, v in sorted(by.items()))
        print(f"   {h['name']:<4} {n:>4} items, {len(h['categories']):>2} categories   {split}")
        print(f"        {h['canteen']} · {h['phone']}"
              + (f" · {h['hours']}" if h["hours"] else "")
              + f" · {len(h['pages'])} scanned page(s)")

    if warnings:
        print(f"\n{len(warnings)} item(s) with no usable Diet — shown to everyone "
              f"and marked unknown, never assumed vegetarian:")
        for w in warnings:
            print(f"   {w}")
