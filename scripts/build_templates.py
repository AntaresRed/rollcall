#!/usr/bin/env python3
"""
Write templates/ — a blank, correctly shaped example of every file the
builders read.

    python3 scripts/build_templates.py

These are not data. They are the answer to "what exactly do you need from
me?", handed to whoever supplies each sheet — the mess secretary, the
placement committee, the academic office — so that the file comes back in a
shape the builder already understands.

Generated rather than hand-made on purpose. A template is a promise about
what a parser accepts, and a promise nobody can regenerate goes stale the
first time the parser changes. Every rule written into these sheets is one
that scripts/build_*.py actually enforces; when a builder changes, change it
here too and re-run.

Each workbook ships with a README.md beside it. The guidance is deliberately
NOT a sheet inside the workbook: several builders read every sheet in the
book as data, so an explanatory tab is indistinguishable from a mess that has
been misfiled — the first draft of these templates was rejected by its own
parser for exactly that reason.
"""

import csv
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = "templates"

HEAD_FILL = PatternFill("solid", fgColor="11151A")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
NOTE_FONT = Font(size=10)
TITLE_FONT = Font(bold=True, size=13)
RULE_FONT = Font(bold=True, size=10)
EG_FONT = Font(italic=True, color="666E77", size=10)


def sheet(wb, title, headers, rows, widths=None, note=None):
    """One data sheet, and nothing else in it.

    No heading rows, no note rows, no explanatory tab. Several builders read
    the first non-empty row as the header and every sheet in the book as
    data, so anything helpful added inside the workbook is indistinguishable
    from a mess that has been misfiled. Guidance lives in the README beside
    the file, where it can also be read without opening Excel.

    `note` is accepted and ignored, so the call sites can keep saying what a
    sheet is for; it goes into the README instead.
    """
    ws = wb.create_sheet(title)
    start = 1

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start, column=i, value=h)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
    for r, row in enumerate(rows, start=start + 1):
        for i, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    return ws


def guide(wb, title, lines, name=None):
    """The workbook's instructions, written beside it rather than inside it.

    `wb` is unused and kept only so the callers read as "this workbook, and
    this is what it means".
    """
    body = [f"# {title}", ""]
    for line in lines:
        body.append(f"**{line}**" if line.endswith(":") else line)
        body.append("")
    path = os.path.join(OUT, f"{name}.README.md")
    with open(path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(body).rstrip() + "\n")
    print(f"  {path}")


def save(wb, name):
    wb.remove(wb["Sheet"])
    path = os.path.join(OUT, name)
    wb.save(path)
    print(f"  {path}")


# ---------------------------------------------------------------- day mess

def day_mess():
    wb = openpyxl.Workbook()
    guide(wb, "Day Mess Menu — one sheet per hostel", [
        "Read by: scripts/build_menu.py    Feeds: the Day mess tab",
        "",
        "Sheet names:",
        "One sheet per hostel. The hostel is read from the FIRST WORD of the "
        "sheet name, so 'NH Day Mess Menu' and 'NH Menu' both work, but two "
        "sheets starting with the same word will stop the build.",
        "Adding a fifth hostel is a new sheet. It needs no code change.",
        "",
        "Columns:",
        "Breakfast, Lunch, Snacks and Dinner must all be present. They are "
        "found by name, so you may reorder them or add columns of your own "
        "in between — but renaming one stops the build.",
        "",
        "Rows:",
        "All seven days, spelled in full, Monday first through Sunday. A "
        "missing or misspelled day stops the build.",
        "Blank spacer rows and stray notes are ignored.",
        "",
        "Items served every day:",
        "A row containing 'Everyday offering' is treated as food served "
        "regardless of the weekday, and shown under every meal. Write it as "
        "'Breakfast: ... Lunch: ...' to split it per meal; anything else is "
        "kept whole and shown as one note.",
        "",
        "What the app does with blanks:",
        "An empty meal cell is allowed and is reported as a warning when the "
        "build runs. It will show as an empty meal in the app.",
    ], name="Day Mess Menu.template.xlsx")

    days = ["Monday", "Tuesday", "Wednesday", "Thursday",
            "Friday", "Saturday", "Sunday"]
    example = {
        "Monday": ["Idli, Sambar, Coconut Chutney, Tea, Coffee",
                   "Rajma, Jeera Rice, Roti, Salad, Curd",
                   "Samosa, Tea, Coffee",
                   "Paneer Butter Masala, Roti, Rice, Gulab Jamun"],
    }
    for tag in ["NH", "OH", "LVH", "WH"]:
        rows = [[d] + example.get(d, ["", "", "", ""]) for d in days]
        if tag == "OH":
            rows.append(["Everyday offering---->",
                         "Breakfast: Bread, Butter, Jam, Boiled Eggs "
                         "Lunch: Salad, Curd, Plain Rice", "", "", ""])
        sheet(wb, f"{tag} Day Mess Menu",
              ["Day", "Breakfast", "Lunch", "Snacks", "Dinner"],
              rows, widths=[22, 34, 34, 26, 34],
              note="Monday is filled in as an example — replace it and the "
                   "six blank days below.")
    save(wb, "Day Mess Menu.template.xlsx")


# -------------------------------------------------------------- night mess

def night_mess():
    wb = openpyxl.Workbook()
    guide(wb, "Night Mess Menu — an Info sheet plus one menu sheet per canteen", [
        "Read by: scripts/build_night_menu.py    Feeds: the Night canteen tab",
        "",
        "The Info sheet:",
        "One row per canteen. 'Hostel' must match the first word of that "
        "canteen's menu sheet name. Column headings are matched loosely, so "
        "'Phone number' still counts as 'Phone'.",
        "Phone: ten digits. Two numbers may be written as '9876543210 / "
        "9123456780' — the app calls and messages the FIRST one, so put the "
        "ordering line first.",
        "Every canteen needs its own number. Two canteens sharing one is "
        "caught by the test suite, because the screen looks perfectly right "
        "while orders walk to the wrong hostel.",
        "Room service (Rs): the flat delivery charge. Leave blank for none.",
        "",
        "The menu sheets:",
        "Named '<HOSTEL> Night Menu'. The four column headings — Category, "
        "Item, Price, Diet — must be spelled exactly. Renaming one stops the "
        "build and prints the header it actually read.",
        "",
        "Diet — only these three words:",
        "veg        no meat, no egg",
        "egg        contains egg but no meat",
        "non-veg    contains meat or fish",
        "Anything else, including a blank, is carried through as "
        "'unconfirmed'. Those items are shown under EVERY filter and marked "
        "with a hollow dot, and the build lists them at the end so they can "
        "be chased. Nothing is ever assumed vegetarian.",
        "",
        "Price:",
        "A whole number. A dish sold in two sizes may be written '22/25'; "
        "the basket adds up using the first figure.",
        "",
        "Categories:",
        "Keep each category's rows together. The app builds its sections in "
        "the order categories first appear, so a category split across two "
        "blocks becomes two sections with the same name.",
        "",
        "The photographed menu (optional, but do send it):",
        "Photograph the menu on the wall and send the images alongside. They "
        "go in public/menu/night/ named <hostel>-1.jpg, <hostel>-2.jpg and "
        "so on — lowercase. The app discovers them from the folder, so no "
        "spreadsheet change is needed. A canteen with no photos simply shows "
        "no 'See Original Menu' button.",
    ], name="Night Mess Menu.template.xlsx")

    sheet(wb, "Info",
          ["Hostel", "Canteen", "Phone", "Hours", "Room service (Rs)"],
          [["NH", "New Hostel Night Canteen", "9876543210",
            "11:00 PM - 4:00 AM", 5],
           ["WH", "W.H Night Cafeteria (Tagore)", "9123456780 / 9988776655",
            "11:00 PM - 4:00 AM", 5]],
          widths=[10, 34, 26, 22, 18],
          note="One row per canteen. Hostel must match the menu sheet name.")

    for tag in ["NH", "WH"]:
        sheet(wb, f"{tag} Night Menu",
              ["Category", "Item", "Price", "Diet"],
              [["Roll", "Veg Roll", 43, "veg"],
               ["Roll", "Egg Roll", 55, "egg"],
               ["Roll", "Chicken Roll", 56, "non-veg"],
               ["Momos", "Veg Steam Momo", 60, "veg"],
               ["Momos", "Chicken Steam Momo", 100, "non-veg"],
               ["Beverages", "Cold Drink", "22/25", "veg"]],
              widths=[24, 34, 12, 14],
              note="Six example rows — replace them. Keep each category's "
                   "rows together.")
    save(wb, "Night Mess Menu.template.xlsx")


# ------------------------------------------------------------- tuck shops

def tuck():
    wb = openpyxl.Workbook()
    guide(wb, "Tuck Shops — an Info sheet plus one card per shop", [
        "Read by: scripts/build_tuck.py    Feeds: the Tuck shops tab",
        "",
        "The Info sheet:",
        "One row per shop. 'Shop' must match the first word of that shop's "
        "menu sheet name. A shop with no phone number is fine — it is a "
        "counter you walk to — and the screen simply shows no call button "
        "rather than a dead one. Never borrow another shop's number.",
        "",
        "The menu sheets:",
        "Named '<SHOP> Tuck'. Sl No., Item and Price must be spelled exactly. "
        "Diet is optional; see below.",
        "One flat list, no categories — that is how the card on the counter "
        "reads, and inventing sections here would be a shape the shop does "
        "not have.",
        "",
        "Prices, exactly as printed:",
        "Half this card is two prices for one item — '40 / 60' for without "
        "and with cheese, '45 (65)' the same idea in brackets. Type them as "
        "the card does. They are shown as written and never added up; there "
        "is no basket on this screen, so nothing has to decide which of the "
        "two you meant.",
        "",
        "Sl No.:",
        "Kept as printed, gaps and all. The card handed over skips 27, 29-31, "
        "33-34 and 42, and renumbering would break the one thing the column "
        "is for — matching a row against the card on the counter.",
        "",
        "Diet — optional, and blank today:",
        "Fill it with veg, egg or non-veg and the Veg only / No meat filter "
        "appears on the screen by itself. Leave it blank and every item reads "
        "as unconfirmed and is shown to everyone, which is the honest state "
        "when nobody has been through the card. Nothing is ever guessed from "
        "an item's name.",
    ], name="Tuck Shops.template.xlsx")

    sheet(wb, "Info", ["Shop", "Name", "Phone", "Hours"],
          [["MOHANDA", "Mohan Da", "8100294443", ""],
           ["TAGORE", "Tagore Tuck Shop", "", ""]],
          widths=[14, 26, 18, 22])

    for tag in ["MOHANDA", "TAGORE"]:
        sheet(wb, f"{tag} Tuck", ["Sl No.", "Item", "Price", "Diet"],
              [[1, "Plain Cheese Sandwich Grilled", "40", ""],
               [2, "Veg Cheese Sandwich Grilled", "60", ""],
               [13, "Paneer Masala Patty / with cheese", "40 / 60", ""],
               [22, "Double Egg Bread Omelet (Cheese)", "45 (65)", ""],
               [44, "Masala Rice (Veg / Egg / Chicken / Paneer)",
                "50 / 70 / 75 / 75", ""]],
              widths=[10, 52, 20, 12])
    save(wb, "Tuck Shops.template.xlsx")


# --------------------------------------------------------------------- POR

def por():
    wb = openpyxl.Workbook()
    guide(wb, "POR Contacts Sheet — eight sheets, four different shapes", [
        "Read by: scripts/build_por.py    Feeds: the POR Details screen",
        "",
        "Sheet names must match exactly:",
        "Student Council · Preparation Committee · Placement Representatives "
        "· Clubs · SIGs · Chapters · Cultural Cell · Sports Council",
        "",
        "Columns are read by POSITION, not by heading:",
        "The headings are there for the person filling the sheet in. The "
        "parser counts columns, so a column inserted in the middle silently "
        "shifts everything after it. Add new columns at the far right only.",
        "",
        "Phone numbers:",
        "Ten digits starting 6-9. '+91' and spaces are fine and are stripped. "
        "Anything that is not a valid Indian mobile is dropped and reported "
        "when the build runs, so check that list.",
        "",
        "Clubs, SIGs and Chapters:",
        "Fill the Group cell on the FIRST row of each group only; leave it "
        "blank on the rows beneath and they are taken as belonging to the "
        "group above.",
        "",
        "Sports Council:",
        "A row reading 'Sports Captains' on its own starts a second section "
        "within the sheet, and a repeated header row after it is fine.",
        "",
        "Corrections:",
        "Numbers known to be wrong in the sheet are fixed in a list inside "
        "build_por.py, and re-checked on every run. If the workbook comes "
        "back with one already fixed — or changed to something else — the "
        "build stops rather than quietly reapplying a stale correction.",
        "",
        "This template will not build as-is, and that is correct:",
        "Running build_por.py against this blank template fails with "
        "'ADDITIONS no longer describe the workbook', naming real students it "
        "expected to find. That is the guard above doing its job — the "
        "template has none of them in it. A real filled-in workbook builds "
        "cleanly. Nothing to fix here.",
    ], name="POR Contacts Sheet.template.xlsx")

    sheet(wb, "Student Council", ["Post", "Name", "Phone", "Email"],
          [["President", "A Student", "9876543210", "astudent2027@email.iimcal.ac.in"],
           ["General Secretary", "B Student", "9123456780", ""]],
          widths=[30, 26, 16, 38])
    sheet(wb, "Preparation Committee", ["Name", "Contact"],
          [["C Student", "9876543211"]], widths=[26, 16],
          note="No post column on this sheet — names and numbers only.")
    sheet(wb, "Placement Representatives", ["Name", "Contact"],
          [["D Student", "9876543212"]], widths=[26, 16],
          note="A decorative title row above the names is fine; it is skipped.")
    for name in ["Clubs", "SIGs", "Chapters"]:
        sheet(wb, name, ["Group", "Post", "Name", "Email", "Contact"],
              [["Finance Club", "Secretary", "E Student",
                "estudent2027@email.iimcal.ac.in", "9876543213"],
               ["", "Joint Secretary", "F Student", "", "9876543214"],
               ["Consult Club", "Secretary", "G Student", "", "9876543215"]],
              widths=[26, 24, 26, 38, 16],
              note="Leave Group blank to continue the group above.")
    sheet(wb, "Cultural Cell", ["Post", "Name", "Email", "Contact"],
          [["Secretary", "H Student", "", "9876543216"]],
          widths=[26, 26, 38, 16])
    sheet(wb, "Sports Council", ["Vertical", "Name", "Contact"],
          [["Sports Secretary", "I Student", "9876543217"],
           ["Sports Captains", "", ""],
           ["Cricket", "J Student", "9876543218"]],
          widths=[26, 26, 16],
          note="'Sports Captains' alone on a row starts the captains section.")
    save(wb, "POR Contacts Sheet.template.xlsx")


# ------------------------------------------------------- the class schedule

def pgp2_schedule():
    wb = openpyxl.Workbook()
    guide(wb, "Class Schedule (PGP2) — the weekly grid and the term calendar", [
        "Read by: scripts/build_catalogue.py    Feeds: every timetable, every "
        "alert, and the attendance maths",
        "",
        "This is the most important file in the portal. Everything else is "
        "reference; this one decides what each student sees and when they are "
        "told to go to class.",
        "",
        "Sheets:",
        "One whose name STARTS WITH 'Class Schedule' — the weekly grid.",
        "One whose name STARTS WITH 'Calendar' — the term dates.",
        "Optionally, one sheet per block course taught on fixed dates. Where "
        "one exists it overrules the grid, because the grid only shows which "
        "slot the course occupies and the detail sheet says which dates it "
        "actually meets.",
        "",
        "The grid:",
        "Day names down the first column, spelled in full. Time slots across "
        "the top as HH:MM. Slot columns are found by reading the header, so "
        "they may be reordered.",
        "Recognised start times: 08:30, 10:15, 12:00, 14:30, 16:15, 18:00. "
        "Every class is 90 minutes. A start time not on that list stops the "
        "build rather than being guessed at.",
        "",
        "What goes in a grid cell:",
        "Course Name-SECTION (Instructor codes) Venue",
        "e.g.  Marketing Research-A (SKB) L-51",
        "A course running only half the term carries its phase in brackets:",
        "e.g.  Bank Management (Post Mid Term) (RC) Amphi West 100",
        "Leave a cell empty when nothing runs in that slot.",
        "",
        "The calendar sheet:",
        "Two columns — a label on the left, dates on the right. The labels "
        "that are read must contain these words:",
        "duration of term · pre mid term classes · post mid term classes · "
        "mid term exam · summer placement · puja vacation · end term exam",
        "Write dates in full: 'August 4, 2026 to September 20, 2026'.",
        "",
        "A warning about running the build by hand:",
        "build_catalogue.py also writes supabase/repair-stale-classes.sql, at "
        "that fixed path, whatever output file you give it. Running it "
        "against a test workbook overwrites the real one. Commit or stash "
        "before experimenting.",
        "",
        "Changes after the file is issued:",
        "The institute amends schedules and does not reissue the workbook. "
        "Do NOT hand-edit this file for those. They go in data/overrides.json, "
        "which is reapplied on every rebuild and fails loudly if the sheet "
        "later changes underneath it.",
    ], name="Class Schedule Term.template.xlsx")

    slots = ["08:30", "10:15", "12:00", "14:30", "16:15", "18:00"]
    rows = [["Monday", "Marketing Research-A (SKB) L-51", "", "",
             "Bank Management (Post Mid Term) (RC) Amphi West 100", "", ""],
            ["Tuesday", "", "", "", "", "", ""],
            ["Wednesday", "", "", "", "", "", ""],
            ["Thursday", "", "", "", "", "", ""],
            ["Friday", "", "", "", "", "", ""],
            ["Saturday", "", "", "", "", "", ""]]
    sheet(wb, "Class Schedule Term-X", ["Day"] + slots, rows,
          widths=[14] + [30] * len(slots),
          note="Monday shows the two cell formats. Empty slot = no class.")

    sheet(wb, "Calendar (Term-X)", ["", ""], [
        ["Duration of Term", "August 4, 2026 to November 28, 2026"],
        ["Pre Mid Term Classes", "August 4, 2026 to September 20, 2026"],
        ["Mid Term Examination", "September 21, 2026 to September 26, 2026"],
        ["Post Mid Term Classes", "September 28, 2026 to November 21, 2026"],
        ["Puja Vacation", "October 8, 2026 to October 14, 2026"],
        ["Summer Placement", "November 2, 2026 to November 7, 2026"],
        ["End Term Examination", "November 23, 2026 to November 28, 2026"],
    ], widths=[34, 52],
        note="Label on the left, dates on the right. Months written in full.")
    save(wb, "Class Schedule Term.template.xlsx")


def pgp1_schedule():
    wb = openpyxl.Workbook()
    guide(wb, "PGP1 Schedule — a core curriculum, so the section decides everything", [
        "Read by: scripts/build_pgp1_catalogue.py    Feeds: the first-year "
        "timetable",
        "",
        "Why this is a separate file from the PGP2 schedule:",
        "First years do not pick courses. Every student is in one section, "
        "A to F, and the whole timetable follows from it. So each grid cell "
        "names its section, and the app shows a section picker rather than a "
        "course search.",
        "",
        "Sheets — all four must be present, named exactly:",
        "Pre Mid Term-II · Post Mid Term-II · Instructors · Calendar",
        "The two grids are the two halves of the term. A course that only "
        "runs in one half appears on one sheet, which is how the app knows "
        "not to alert for it in the other half.",
        "",
        "The grids:",
        "The header row must contain a cell reading 'Dayslot / Timeslot' — "
        "that cell is how the parser finds which column holds the day names, "
        "so the column may move but that heading may not disappear.",
        "Day names go under it, spelled in full. Time slots across the top as "
        "HH:MM, found by reading the same header row, which may sit anywhere "
        "in the first eleven rows.",
        "Recognised start times: 08:30, 10:15, 12:00, 14:30, 16:15. First "
        "years have no 18:00 slot.",
        "",
        "What goes in a grid cell:",
        "Course Name (SECTION) then, on a new line, (Instructor code) Room",
        "e.g.  Organizational Behaviour-II (A)",
        "      (VJ) L-21",
        "Every course must cover sections A to F exactly once across the "
        "grid. A section listed twice, or missing, stops the build — that is "
        "the check that catches a copy-paste slip in a sheet where every row "
        "looks alike.",
        "",
        "Rooms follow the section, not the course:",
        "A=L-21, B=L-22, C=L-31, D=L-32, E=N-31, F=N-32 in a typical term. "
        "Write the room in the cell anyway; it is read from there, not "
        "assumed.",
        "",
        "The Instructors sheet:",
        "Course Code in column B, Course Name in C, Credits in D, "
        "Instructors in E. The header row may sit anywhere in the first "
        "seven rows as long as one cell reads 'Course Code'.",
        "List several instructors on separate lines within the cell. Mark "
        "who teaches which sections in brackets — 'Secs-A,C', 'Sec-B', "
        "'(A, B and C)' and 'Secs.CDEF' are all understood.",
        "Name the sections explicitly. Between them the instructors must "
        "cover A to F exactly once, and a phrase like 'all sections' does not "
        "count — the build stops rather than assuming which six it meant.",
        "Course names here must match the grid's spelling closely; where "
        "they have drifted the build says so rather than guessing.",
        "",
        "The Calendar sheet:",
        "Same shape as the PGP2 one — a label on the left, dates on the "
        "right. First-year term dates are genuinely different from the "
        "second years', which is why each cohort has its own.",
        "Every one of these rows is required, and a missing one stops the "
        "build by name: duration of term · pre mid term classes · post mid "
        "term classes · mid term exam · summer placement · puja vacation · end term exam.",
    ], name="PGP1 Schedule.template.xlsx")

    slots = ["08:30", "10:15", "12:00", "14:30", "16:15"]
    rooms = {"A": "L-21", "B": "L-22", "C": "L-31",
             "D": "L-32", "E": "N-31", "F": "N-32"}

    # A course must cover A-F exactly once across its own grid, so the example
    # shows all six. Each course sits on one sheet only, which is how a
    # half-term course tells the app not to alert for it in the other half.
    for name, course, code in [
        ("Pre Mid Term-II", "Organizational Behaviour-II", "VJ"),
        ("Post Mid Term-II", "Quantitative Methods-II", "SR"),
    ]:
        def at(sec, course=course, code=code):
            return f"{course} ({sec})\n({code}) {rooms[sec]}"
        rows = [["", "Monday", at("A"), at("B"), at("C"), "", ""],
                ["", "Tuesday", at("D"), at("E"), at("F"), "", ""],
                ["", "Wednesday", "", "", "", "", ""],
                ["", "Thursday", "", "", "", "", ""],
                ["", "Friday", "", "", "", "", ""],
                ["", "Saturday", "", "", "", "", ""]]
        sheet(wb, name, ["", "Dayslot / Timeslot"] + slots, rows,
              widths=[6, 20] + [30] * len(slots),
              note="Day names go under the Dayslot / Timeslot header.")

    sheet(wb, "Instructors",
          ["Sl. No.", "Course Code", "Course Name", "Credits", "Instructors"],
          [[1, "OB-II", "Organizational Behaviour-II", 3.0,
            "Prof. V Jain (Secs-A,B,C)\nProf. R Kumar (Secs-D,E,F)"],
           [2, "QM-II", "Quantitative Methods-II", 1.5,
            "Prof. S Roy (Secs-A,B,C)\nProf. T Bose (Secs-D,E,F)"]],
          widths=[10, 18, 40, 12, 46],
          note="Course Code in column B. Sections in brackets after each name.")

    sheet(wb, "Calendar", ["", ""], [
        ["Duration of Term", "August 4, 2026 to November 28, 2026"],
        ["Pre Mid Term Classes", "August 4, 2026 to September 20, 2026"],
        ["Mid Term Examination", "September 21, 2026 to September 26, 2026"],
        ["Post Mid Term Classes", "September 28, 2026 to November 21, 2026"],
        ["Puja Vacation", "October 8, 2026 to October 14, 2026"],
        ["Summer Placement", "November 2, 2026 to November 7, 2026"],
        ["End Term Examination", "November 23, 2026 to November 28, 2026"],
    ], widths=[34, 52])
    save(wb, "PGP1 Schedule.template.xlsx")


# ------------------------------------------------------------- the two TSVs

def tsv(name, headers, rows, notes):
    path = os.path.join(OUT, name)
    with open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter="\t")
        w.writerow(headers)
        for row in rows:
            w.writerow(row)
    with open(os.path.join(OUT, name.replace(".tsv", ".README.md")),
              "w", encoding="utf-8", newline="\n") as fh:
        fh.write(notes)
    print(f"  {path}")


def faculty_directory():
    tsv(
        "FacultyDirectory.template.tsv",
        ["Name", "Room", "Extension", "Direct", "Email"],
        [["Abhipsa Pal", "K-208", "2080", "033-7121-2080", "abhipsapal@iimcal.ac.in"],
         ["Manish Thakur (Dean NIER)", "P-303", "1120", "033-7121-1120",
          "dean_nier@iimcal.ac.in"],
         ["Manish Thakur", "B-308", "2120", "033-7121-2120", "mt@iimcal.ac.in"]],
        """# FacultyDirectory.tsv

**Read by:** `scripts/build_directory.py` (and `scripts/build_faculty.py`)
**Feeds:** the Faculty directory screen, and the email shown against each course

Tab-separated. The first row is a header and is skipped.

## Columns, in this order

| # | Column | Notes |
|---|---|---|
| 1 | Name | May carry an administrative title in brackets — see below |
| 2 | Room | |
| 3 | Extension | |
| 4 | Direct | Full outside line |
| 5 | Email | |

**Columns are read by position, not by heading.** A column inserted in the
middle shifts everything after it. Add new columns at the far right only.

## Two rows for one person

Deans and chairpersons appear twice: once under their own name and office,
and once under the administrative title with that office's room and role
address. Write the title in brackets after the name —

```
Manish Thakur (Dean NIER)	P-303	1120	033-7121-1120	dean_nier@iimcal.ac.in
Manish Thakur	B-308	2120	033-7121-2120	mt@iimcal.ac.in
```

— and the app shows one person holding both offices, their own first. Do not
merge them into one row: each carries a number somebody might need to dial.

## Nothing on file

Write `—`, `-`, `n/a` or leave the cell empty. Any of those is understood as
"not on file" and is left out of the card rather than printed as if it were a
room number.

## A caution

This builder currently has **no failure path**. If the columns move, it will
write a plausible-looking file with the wrong things in the wrong fields, and
say nothing. Check the count it prints at the end against what you expected.
""")


def course_outline():
    tsv(
        "CourseProfOutline.template.tsv",
        ["Sl. No.", "Course Code", "Course Name", "Area", "Credit", "Instructors"],
        [["", "Finance & Control", "", "", "", ""],
         [1, "FC-401", "Corporate Finance", "Finance & Control", "3",
          "Prof. A Banerjee (CC)  Prof. B Sen (VF)"],
         [2, "MK-402", "Marketing Research (Pre Mid Term)", "Marketing", "1.5",
          "Prof. C Das (CC)"]],
        """# CourseProfOutline.tsv

**Read by:** `scripts/build_faculty.py`
**Feeds:** the professor shown against each course

Tab-separated. Everything above the header row is ignored, and the header row
is found by looking for a first cell reading exactly **`Sl. No.`** — so that
cell has to be spelled that way.

## Columns, in this order

| # | Column | Notes |
|---|---|---|
| 1 | Sl. No. | Also used to find the header row |
| 2 | Course Code | Required — a row without one is skipped |
| 3 | Course Name | Required |
| 4 | Area | Read but not used |
| 5 | Credit | `3` or `1.5` |
| 6 | Instructors | See below |

**Columns are read by position, not by heading.**

## Group heading rows

A row with a group name and no course code — `Finance & Control` on its own —
is skipped. You do not need to remove them.

## The instructors cell

Titles are required, because they are how several names run together in one
cell are told apart:

```
Prof. Sudhir Jaiswall (CC)      Mr. Balachandran R (VF)
```

`Prof.`, `Dr.`, `Mr.`, `Ms.` and `Mrs.` are all recognised. `(CC)` marks the
course coordinator and `(VF)` a visiting faculty member.

Names are matched against `FacultyDirectory.tsv` to find each professor's
email. Matching is deliberately conservative — a close-but-not-certain match
is left unmatched rather than risking the wrong person's address on a course.
Visiting faculty usually have no institute address at all, and coming back
unmatched is the correct outcome for them.

The build prints every internal professor it could not match. That list is
worth reading: it is usually a spelling difference between this sheet and the
directory.

## A caution

This builder currently has **no failure path**. If the columns move it will
write a file with the wrong things in the wrong fields and say nothing. Check
the counts it prints at the end.
""")


def index():
    with open(os.path.join(OUT, "README.md"), "w",
              encoding="utf-8", newline="\n") as fh:
        fh.write("""# Input templates

A blank, correctly shaped example of every file the portal is built from.
Hand the relevant one to whoever supplies that data.

**These are examples, not data.** Nothing here is read by the app or by any
build. Fill a copy in, put it wherever you keep the real sheets, and point
the builder at that.

| Template | Given by | Read by | Feeds |
|---|---|---|---|
| `Class Schedule Term.template.xlsx` | Academic office | `build_catalogue.py` | Second-year timetable, alerts, attendance |
| `PGP1 Schedule.template.xlsx` | Academic office | `build_pgp1_catalogue.py` | First-year timetable |
| `CourseProfOutline.template.tsv` | Academic office | `build_faculty.py` | Professor shown per course |
| `FacultyDirectory.template.tsv` | Institute directory | `build_directory.py` | Faculty directory screen |
| `POR Contacts Sheet.template.xlsx` | Student Council | `build_por.py` | POR Details screen |
| `Day Mess Menu.template.xlsx` | Mess secretary | `build_menu.py` | Day mess tab |
| `Tuck Shops.template.xlsx` | Tuck shop / mess secretary | `build_tuck.py` | Tuck shops tab |
| `Night Mess Menu.template.xlsx` | Canteen / mess secretary | `build_night_menu.py` | Night canteen tab |

Every template has a `.README.md` beside it saying what each sheet and column
means and what will stop the build.

The guidance is deliberately not a tab inside the workbook. `build_menu.py`
treats every sheet as a hostel and `build_night_menu.py` every sheet but
`Info` as a menu, so an explanatory tab reads as data and stops the build.
That is worth knowing when you fill one in: **do not leave a "Notes" or
"Sheet1" tab in the mess workbooks.**

## Rules that hold for all of them

**Send `.xlsx`.** The builders use openpyxl, which reads `.xlsx` only. A
`.xls`, a Google Sheet or a PDF needs *File → Download as .xlsx* first.

**A renamed column usually stops the build**, and prints the header it
actually read. That is the intended behaviour: a build that stops costs an
hour, and a build that guesses costs a term of wrong alerts.

**Two files are read by column position rather than by heading** — both
`.tsv`s. Inserting a column in the middle of those shifts everything after it
silently. Add columns at the right-hand end only.

**Corrections after the fact do not belong in these files.** The institute
amends schedules without reissuing the workbook. Those go in
`data/overrides.json` (schedules) or the correction lists inside
`build_por.py` (contacts), both of which are reapplied on every rebuild and
fail loudly if the source later changes underneath them.

**Read what the build prints.** Every builder ends with a summary — how many
courses, people or items it found, and what it could not resolve. That
summary is the check that catches a file which parsed cleanly but is not the
file you meant to hand it.

## These have been run through their own parsers

Every template here was fed to the builder that reads it, and the output
checked. A template that its own parser rejects is a wrong template, and the
first draft of these was rejected — the explanatory tab inside each workbook
was being read as a hostel.

One exception, and it is correct behaviour: **`build_por.py` fails against
the blank POR template**, with `ADDITIONS no longer describe the workbook`.
The corrections list inside that builder names real students, and this
template contains none of them. A real filled-in workbook builds cleanly.

## One hazard when testing a build by hand

`build_catalogue.py` also writes `supabase/repair-stale-classes.sql`, at that
fixed path, whatever output file you pass it. Running it against a test
workbook overwrites the real one. Commit or stash before experimenting.

## Regenerating

    python3 scripts/build_templates.py

The templates are generated, so that when a parser changes the template can
change with it. If you edit a builder's expectations, edit
`scripts/build_templates.py` too and re-run.
""")
    print(f"  {os.path.join(OUT, 'README.md')}")


if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    print(f"writing templates into {OUT}/")
    day_mess()
    tuck()
    night_mess()
    por()
    pgp2_schedule()
    pgp1_schedule()
    faculty_directory()
    course_outline()
    index()
    print("done.")
