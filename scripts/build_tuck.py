#!/usr/bin/env python3
"""
Build src/data/tuck.json — the tuck shops.

    python3 scripts/build_tuck.py \
        "data/Tuck Shops.xlsx" src/data/tuck.json

A third sibling of build_menu.py and build_night_menu.py. The day mess is a
week, the night canteens are priced lists by category, and a tuck shop is one
flat priced list you read standing at the counter — close enough to the night
menu to share a screen's furniture, different enough that flattening the three
into one parser would mean a change for one breaking the others.

Sheets:
  Info            Shop | Name | Phone | Hours
  <TAG> Tuck      Sl No. | Item | Price | Diet

Two things worth knowing about the data:

  * Prices are kept exactly as printed. Half this card is written "40 / 60"
    or "45 (65)" — one item at two prices, with and without cheese — and
    picking one of those numbers to store would be inventing a fact. There is
    no basket here, so nothing needs to add them up.
  * `Diet` is optional and currently blank. The app would rather mark an item
    unconfirmed than guess "Veg" from a name, because the times a guess is
    wrong are the times a vegetarian eats meat. Fill the column in and the
    filter starts working; leave it and every item is shown to everyone.
"""

import glob
import json
import os
import re
import sys

import openpyxl

DIETS = {"veg", "egg", "non-veg"}
COLUMNS = ["Sl No.", "Item", "Price"]


def norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def die(msg):
    raise SystemExit(f"tuck: {msg}")


def read_info(ws):
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(r)]
    if not rows:
        die("Info sheet is empty")
    head = [c.lower() for c in rows[0]]

    def col(*names):
        for i, h in enumerate(head):
            if any(n in h for n in names):
                return i
        return None

    ci = {k: col(*n) for k, n in {
        "id": ("shop", "hostel"), "name": ("name",),
        "phone": ("phone",), "hours": ("hours",), "upi": ("upi", "vpa"),
    }.items()}
    if ci["id"] is None:
        die("Info sheet has no Shop column")

    out = {}
    for r in rows[1:]:
        tag = r[ci["id"]].upper()
        if not tag:
            continue
        get = lambda k: (r[ci[k]] if ci[k] is not None and ci[k] < len(r) else "")
        out[tag] = {"name": get("name") or tag, "phone": get("phone"),
                    "hours": get("hours"), "upi": get("upi")}
    return out


def read_items(ws):
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(r)]
    if not rows:
        die(f"{ws.title}: empty")

    head = [c.lower() for c in rows[0]]
    idx = {}
    for want in COLUMNS:
        for i, h in enumerate(head):
            if h == want.lower():
                idx[want] = i
                break
        else:
            die(f"{ws.title}: no '{want}' column (header reads {rows[0]})")
    # Optional: a card nobody has been through yet simply has no diet on it.
    diet_col = next((i for i, h in enumerate(head) if h == "diet"), None)

    items, unknown = [], []
    for r in rows[1:]:
        get = lambda i: r[i] if i is not None and i < len(r) else ""
        sl, name, price = (get(idx[c]) for c in COLUMNS)
        if not name:
            continue
        if not price:
            die(f"{ws.title}: '{name}' has no price")

        diet = get(diet_col).lower()
        if diet not in DIETS:
            if diet:
                unknown.append(f"{name} ({diet})")
            diet = "unknown"

        items.append({"sl": sl, "name": name, "price": price, "diet": diet})

    names = [i["name"] for i in items]
    if len(set(names)) != len(names):
        dupes = sorted({n for n in names if names.count(n) > 1})
        die(f"{ws.title}: the same item twice: {', '.join(dupes)}")

    return items, unknown


def qr_for(qr_dir, tag):
    """The shop's own payment QR, if somebody has put one there.

    Discovered from the directory rather than named in the spreadsheet, the
    same way the night menu finds its photographed pages: adding one is
    dropping `mohanda.png` next to the others.

    Deliberately the shop's OWN image rather than a code generated here from
    the UPI address. Generating one means encoding a payment instruction, and
    a subtly wrong encoding is a payment that goes somewhere subtly wrong. The
    printed code on the counter is the one already known to work.
    """
    for ext in ("png", "jpg", "jpeg", "webp"):
        hit = glob.glob(os.path.join(qr_dir, f"{tag.lower()}.{ext}"))
        if hit:
            return f"/tuck/{os.path.basename(hit[0])}"
    return None


def build(path, qr_dir="public/tuck"):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Info" not in wb.sheetnames:
        die("workbook has no Info sheet")
    info = read_info(wb["Info"])

    shops, warnings = [], []
    for name in wb.sheetnames:
        if name == "Info":
            continue
        tag = norm(name).split()[0].upper()
        items, unknown = read_items(wb[name])
        meta = info.get(tag)
        if meta is None:
            die(f"'{name}' has no matching row in the Info sheet (looked for {tag})")
        if not re.fullmatch(r"[6-9]\d{9}", re.sub(r"\D", "", meta["phone"] or "")):
            # Not fatal — a counter you walk to is still worth listing — but
            # said out loud, because the screen quietly drops the call button.
            warnings.append(f"{tag}: no usable phone number, so no call button")
        upi = (meta.get("upi") or "").strip()
        if upi and not re.fullmatch(r"[a-z0-9.\-_]{2,}@[a-z]{2,}", upi, re.I):
            # Never nearly-right: a malformed address sends money nowhere, and
            # the student only finds out standing at the counter.
            die(f"{tag}: '{upi}' is not a UPI address (it should look like "
                f"name@bank)")
        if not upi:
            warnings.append(f"{tag}: no UPI address, so no pay button")
        qr = qr_for(qr_dir, tag)
        if not upi and not qr:
            warnings.append(f"{tag}: no UPI address and no QR, so no way to pay "
                            f"from the app")
        shops.append({"id": tag.lower(), "tag": tag, **meta,
                      "qr": qr, "items": items})
        warnings += [f"{tag}: {u}" for u in unknown]

    if not shops:
        die("no tuck shop sheets found")

    # Two shops printing the same card is a real thing — worth saying once, so
    # that a copy-paste mistake and a genuinely shared menu look different.
    seen = {}
    for s in shops:
        key = tuple(sorted(i["name"] for i in s["items"]))
        seen.setdefault(key, []).append(s["tag"])
    for tags in seen.values():
        if len(tags) > 1:
            warnings.append(f"identical item lists: {', '.join(tags)}")

    return {"shops": shops}, warnings


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "data/Tuck Shops.xlsx"
    dest = sys.argv[2] if len(sys.argv) > 2 else "src/data/tuck.json"
    qr_dir = sys.argv[3] if len(sys.argv) > 3 else "public/tuck"

    data, warnings = build(src, qr_dir)
    with open(dest, "w", encoding="utf-8", newline="\n") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)

    total = sum(len(s["items"]) for s in data["shops"])
    print(f"{len(data['shops'])} tuck shops, {total} items -> {dest}")
    for s in data["shops"]:
        by = {}
        for i in s["items"]:
            by[i["diet"]] = by.get(i["diet"], 0) + 1
        split = "  ".join(f"{k} {v}" for k, v in sorted(by.items()))
        print(f"   {s['tag']:<8} {len(s['items']):>3} items   {split}")
        print(f"        {s['name']} · {s['phone'] or 'no number'}"
              + (f" · {s['hours']}" if s["hours"] else ""))

    if warnings:
        print(f"\n{len(warnings)} thing(s) to look at:")
        for w in warnings:
            print(f"   {w}")
