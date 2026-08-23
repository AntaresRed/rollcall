#!/usr/bin/env python3
"""
Build RollCall's course catalogue from the official IIM Calcutta Term V
spreadsheet.

    python3 scripts/build_catalogue.py \
        data/Class_Schedule_Term-V_AY-2026-27.xlsx \
        src/data/catalogue.json \
        data/overrides.json

The workbook has three kinds of sheet:

  "Class Schedule Term-V"  the weekly grid: day rows x time-slot columns
  "Calendar (Term-V)"      term dates and the periods with no classes
  per-course sheets        courses that run on fixed dates instead of weekly,
                           because visiting faculty teach in blocks

A per-course sheet wins over the grid: the grid shows which slots a block
course occupies, the detail sheet says which dates it actually meets.
"""

import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime

import openpyxl

DAYS = {
    "MONDAY": 1, "TUESDAY": 2, "WEDNESDAY": 3, "THURSDAY": 4,
    "FRIDAY": 5, "SATURDAY": 6, "SUNDAY": 7,
}
DAY_NAMES = {1: "Mon", 2: "Tue", 3: "Wed", 4: "Thu", 5: "Fri", 6: "Sat", 7: "Sun"}

SLOT_END = {
    "08:30": "09:45", "10:15": "11:30", "12:00": "13:15",
    "14:30": "15:45", "16:15": "17:30", "18:00": "19:15",
}

# 3 credits = 20 classes at 75%; 1.5 credits = 10 classes at 80%.
CREDIT_RULES = {
    3.0: {"total_classes": 20, "min_pct": 75},
    1.5: {"total_classes": 10, "min_pct": 80},
}

GRID_SHEET = "Class Schedule"
CALENDAR_SHEET = "Calendar"

PHASE_RE = re.compile(r"\(\s*(pre|post)[-\s]?mid\s*term\s*\)", re.I)
SECTION_RE = re.compile(r"-([AB])\s*$")
CREDIT_RE = re.compile(r"([\d.]+)\s*(?:Cr|Credit)", re.I)
TIME_RE = re.compile(
    r"(\d{1,2}):(\d{2})\s*(?:(am|pm))?\s*-\s*(\d{1,2}):(\d{2})\s*(am|pm)?", re.I
)


# ---------------------------------------------------------------- helpers

def norm(text) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def acronym(name: str, taken: set) -> str:
    """A short, stable, guessable code. Students see names, not codes, but the
    code is what links a saved timetable back to the catalogue."""
    stop = {"and", "of", "in", "for", "the", "a", "an", "to", "on", "with",
            "from", "its", "at", "by", "as"}
    words = [w for w in re.findall(r"[A-Za-z0-9]+", name) if w.lower() not in stop]
    code = "".join(w[0] for w in words).upper()[:6] or "X"
    if code not in taken:
        taken.add(code)
        return code
    for n in range(2, 100):
        alt = f"{code[:5]}{n}"
        if alt not in taken:
            taken.add(alt)
            return alt
    raise SystemExit(f"could not derive a unique code for {name!r}")


def to_24h(hour, minute, meridiem):
    if meridiem:
        m = meridiem.lower()
        if m == "pm" and hour != 12:
            hour += 12
        elif m == "am" and hour == 12:
            hour = 0
    return f"{hour:02d}:{minute:02d}"


def parse_time_range(text):
    """'4:15 - 5:45 pm' -> ('16:15', '17:45'). The meridiem is usually written
    once, at the end, and applies to both halves."""
    m = TIME_RE.search(norm(text))
    if not m:
        return None
    sh, sm, s_mer, eh, em, e_mer = m.groups()
    mer = e_mer or s_mer
    start = to_24h(int(sh), int(sm), s_mer or mer)
    end = to_24h(int(eh), int(em), mer)
    if start > end:  # '12:00 - 1:30 pm' would otherwise read 12:00 -> 01:30
        start = to_24h(int(sh) % 12, int(sm), "am" if mer == "pm" else "pm")
    return start, end


def parse_cell(text):
    """Split a grid cell into (name, section, phase, venue).

    Cells look like:
        'Course Name-A (Instructor codes) Venue'
        'Course Name (Pre Mid Term) (Instructors) Venue'
    """
    raw = norm(text)
    if not raw:
        return None

    phase = "full"
    found = PHASE_RE.search(raw)
    if found:
        phase = "pre_mid" if found.group(1).lower() == "pre" else "post_mid"
        raw = norm(PHASE_RE.sub(" ", raw))

    open_idx = raw.find("(")
    if open_idx == -1:
        name, venue = raw, ""
    else:
        name = raw[:open_idx]
        depth, close_idx = 0, None
        for i in range(open_idx, len(raw)):
            if raw[i] == "(":
                depth += 1
            elif raw[i] == ")":
                depth -= 1
                if depth == 0:
                    close_idx = i
                    break
        venue = raw[close_idx + 1:] if close_idx is not None else ""

    name = norm(name)
    section = "A"
    sec = SECTION_RE.search(name)
    if sec:
        section = sec.group(1)
        name = norm(SECTION_RE.sub("", name))

    return name, section, phase, norm(venue).strip(" ,/")


# ---------------------------------------------------------------- sheets

def parse_grid(ws):
    slots, header_row = [], None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [norm(c) for c in row]
        if any(re.fullmatch(r"\d{2}:\d{2}", c) for c in cells):
            header_row, slots = i, cells
            break
    if header_row is None:
        raise SystemExit("no slot header row found in the grid sheet")

    entries, current_day = [], None
    for row in ws.iter_rows(min_row=header_row, values_only=True):
        cells = [norm(c) for c in row]
        joined = " ".join(cells).lower()

        # Below the grid sits a legend of instructor abbreviations and a
        # signature block. Everything from there on is not a class.
        if "abbreviation" in joined or "chairperson" in joined or joined.startswith("dated"):
            break

        label = cells[0].upper() if cells else ""
        if label in DAYS:
            current_day = DAYS[label]
        if any(re.fullmatch(r"\d{2}:\d{2}", c) for c in cells) and label not in DAYS:
            continue  # a repeated header mid-sheet
        if current_day is None:
            continue

        for col, cell in enumerate(cells):
            if not cell or col >= len(slots):
                continue
            slot = slots[col]
            if not re.fullmatch(r"\d{2}:\d{2}", slot):
                continue
            parsed = parse_cell(cell)
            if not parsed or not parsed[0]:
                continue
            name, section, phase, venue = parsed
            entries.append({"name": name, "section": section, "phase": phase,
                            "venue": venue, "day": current_day, "start": slot,
                            "end": SLOT_END.get(slot, slot)})
    return entries


def parse_calendar(ws):
    text = {}
    for row in ws.iter_rows(values_only=True):
        cells = [norm(c) for c in row]
        if len(cells) >= 2 and cells[0] and cells[1]:
            text[cells[0].lower()] = cells[1]

    def span(fragment):
        for k, v in text.items():
            if fragment in k:
                return v
        return ""

    def dates_in(s):
        year = re.search(r"(20\d{2})", s)
        year = int(year.group(1)) if year else date.today().year
        out = []
        for month_name, day in re.findall(r"([A-Z][a-z]+)\s+(\d{1,2})", s):
            try:
                out.append(datetime.strptime(
                    f"{month_name} {day} {year}", "%B %d %Y").date().isoformat())
            except ValueError:
                continue
        return out

    breaks = []
    for label, frag in [("Mid-term exams", "mid term exam"),
                        ("Summer placement", "summer placement"),
                        ("Puja vacation", "puja vacation"),
                        ("End-term exams", "end term exam")]:
        d = dates_in(span(frag))
        if len(d) >= 2:
            breaks.append({"label": label, "from": d[0], "to": d[1],
                           "note": norm(span(frag))})

    duration = dates_in(span("duration of term"))
    pre = dates_in(span("pre mid term classes"))
    post = dates_in(span("post mid term classes"))

    return {
        "term_start": pre[0] if pre else (duration[0] if duration else None),
        "pre_mid_end": pre[1] if len(pre) > 1 else None,
        "post_mid_start": post[0] if post else None,
        "term_end": post[1] if len(post) > 1 else (duration[1] if len(duration) > 1 else None),
        "breaks": breaks,
    }


def parse_course_sheet(ws):
    header, code, credits = None, None, 3.0
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            s = norm(cell)
            if not s:
                continue
            m = CREDIT_RE.search(s)
            if m and "," in s:
                code, _, rest = s.partition(",")
                code = norm(code)
                credits = float(m.group(1))
                header = norm(CREDIT_RE.sub("", rest)).strip(" -–—")
    if not header:
        return None

    venue = ""
    for row in ws.iter_rows(values_only=True):
        cells = [norm(c) for c in row]
        if any(c.lower() == "venue" for c in cells):
            venue = next((c for c in cells if c and c.lower() != "venue"), "")
            break

    sessions, current_date = [], None
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, datetime):
                current_date = cell.date().isoformat()
            elif isinstance(cell, date):
                current_date = cell.isoformat()
        times = None
        for cell in row:
            got = parse_time_range(norm(cell)) if cell else None
            if got:
                times = got
                break
        if times and current_date:
            sessions.append({"date": current_date, "start": times[0], "end": times[1]})

    if not sessions:
        return None

    return {"official_code": code, "name": header, "credits": credits,
            "venue": venue,
            "sessions": sorted(sessions, key=lambda s: (s["date"], s["start"]))}


# ---------------------------------------------------------------- assembly

def build(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    grid_name = next(s for s in wb.sheetnames if s.startswith(GRID_SHEET))
    cal_name = next(s for s in wb.sheetnames if s.startswith(CALENDAR_SHEET))
    detail_names = [s for s in wb.sheetnames if s not in (grid_name, cal_name)]

    entries = parse_grid(wb[grid_name])
    calendar = parse_calendar(wb[cal_name])
    details = [d for d in (parse_course_sheet(wb[n]) for n in detail_names) if d]

    taken, by_name = set(), {}
    for e in entries:
        course = by_name.get(e["name"])
        if not course:
            credits = 1.5 if e["phase"] in ("pre_mid", "post_mid") else 3.0
            course = {
                "code": acronym(e["name"], taken),
                "name": e["name"],
                "phase": e["phase"],
                "credits": credits,
                **CREDIT_RULES[credits],
                "venue": e["venue"],
                "schedule_type": "weekly",
                "sections": defaultdict(list),
            }
            by_name[e["name"]] = course
        meeting = {"day": e["day"], "start": e["start"], "end": e["end"]}
        if meeting not in course["sections"][e["section"]]:
            course["sections"][e["section"]].append(meeting)

    matched = []
    for d in details:
        target = d["name"].lower()
        course = None
        for name, c in by_name.items():
            if name.lower() == target or target.startswith(name.lower()[:28]) \
               or name.lower().startswith(target[:28]):
                course = c
                break
        if not course:
            course = {
                "code": acronym(d["name"], taken), "name": d["name"],
                "phase": "full", "credits": d["credits"],
                **CREDIT_RULES[d["credits"]], "venue": d["venue"],
                "schedule_type": "weekly", "sections": defaultdict(list),
            }
            by_name[d["name"]] = course

        course["credits"] = d["credits"]
        course.update(CREDIT_RULES[d["credits"]])
        course["official_code"] = d["official_code"]
        course["schedule_type"] = "dates"
        course["venue"] = d["venue"] or course.get("venue", "")
        course["sections"] = {
            "A": [{"date": s["date"],
                   "day": date.fromisoformat(s["date"]).isoweekday(),
                   "start": s["start"], "end": s["end"]} for s in d["sessions"]]
        }
        matched.append((course["code"], d["official_code"], len(d["sessions"])))

    courses = []
    for c in by_name.values():
        c["sections"] = {
            k: sorted(v, key=lambda m: (m.get("date", ""), m.get("day", 0), m["start"]))
            for k, v in sorted(dict(c["sections"]).items())
        }
        courses.append(c)

    return ({"term": "TERM-V: AY 2026-2027", "slots": list(SLOT_END.keys()),
             "calendar": calendar,
             "courses": sorted(courses, key=lambda c: c["name"].lower())},
            matched)


# ---------------------------------------------------------------- overrides

def apply_overrides(catalogue, overrides_path):
    try:
        with open(overrides_path, encoding="utf-8") as fh:
            spec = json.load(fh)
    except FileNotFoundError:
        return []

    by_code = {c["code"]: c for c in catalogue["courses"]}
    by_name = {c["name"].lower(): c for c in catalogue["courses"]}
    applied = []

    def resolve(ov):
        """Codes are derived from the sheet, so a name is the more durable
        way for an override to point at a course."""
        if ov.get("code") and ov["code"] in by_code:
            return by_code[ov["code"]]
        want = norm(ov.get("name", "")).lower()
        if want:
            if want in by_name:
                return by_name[want]
            for name, c in by_name.items():
                if name.startswith(want[:28]) or want.startswith(name[:28]):
                    return c
        return None

    for ov in spec.get("overrides", []):
        op = ov["op"]

        if op == "add_course":
            sp = ov["to"]
            if sp["code"] in by_code:
                applied.append(f"{sp['code']}: already present")
                continue
            credits = float(sp.get("credits", 3.0))
            rules = dict(CREDIT_RULES[credits])
            # A scratch course for alert testing isn't a real 20-class course.
            for k in ("total_classes", "min_pct"):
                if k in sp:
                    rules[k] = sp[k]
            course = {
                "code": sp["code"], "name": sp["name"],
                "phase": sp.get("phase", "full"), "credits": credits,
                **rules, "venue": sp.get("venue", ""),
                "schedule_type": sp.get("schedule_type", "weekly"),
                "sections": {
                    letter: [
                        {
                            **({"date": m["date"],
                                "day": date.fromisoformat(m["date"]).isoweekday()}
                               if m.get("date") else {"day": m["day"]}),
                            "start": m["start"],
                            "end": m.get("end", SLOT_END.get(m["start"], m["start"])),
                            # Per-meeting phase: a course that runs the whole
                            # term but meets on different days before and
                            # after mid-terms tags each meeting individually,
                            # rather than the whole course being pre_mid or
                            # post_mid. Course-level "phase" stays 'full'.
                            **({"phase": m["phase"]} if m.get("phase") else {}),
                        }
                        for m in meetings
                    ]
                    for letter, meetings in sp["sections"].items()
                },
            }
            catalogue["courses"].append(course)
            by_code[sp["code"]] = course
            by_name[sp["name"].lower()] = course
            applied.append(f"{sp['code']}: {ov.get('note', 'added')}")
            continue

        course = resolve(ov)
        if not course:
            raise SystemExit(
                "override targets a course that isn't in the sheet: "
                f"{ov.get('code') or ov.get('name')!r}"
            )
        letter = ov.get("section", "A")

        if op == "replace_schedule":
            # Wholesale replacement of one section's meetings — for when the
            # grid's pattern is wrong outright rather than off by one slot,
            # so patching meeting-by-meeting would be more fragile than just
            # stating the correct schedule directly.
            meetings_in = ov["to"]["meetings"]
            built = []
            for m in meetings_in:
                if m.get("date"):
                    built.append({
                        "date": m["date"],
                        "day": date.fromisoformat(m["date"]).isoweekday(),
                        "start": m["start"],
                        "end": m.get("end", SLOT_END.get(m["start"], m["start"])),
                    })
                else:
                    built.append({
                        "day": m["day"],
                        "start": m["start"],
                        "end": m.get("end", SLOT_END.get(m["start"], m["start"])),
                        **({"phase": m["phase"]} if m.get("phase") else {}),
                    })
            course["sections"][letter] = sorted(
                built, key=lambda x: (x.get("date", ""), x.get("day", 0), x["start"]))
            course["schedule_type"] = "dates" if any(m.get("date") for m in meetings_in) else "weekly"
            if "venue" in ov["to"]:
                course["venue"] = ov["to"]["venue"]
            applied.append(f"{course['code']}-{letter}: "
                           f"{ov.get('note', 'replace_schedule')} ({len(built)} meetings)")
            continue

        if op == "set_dates":
            sessions = ov["to"]["sessions"]
            course["schedule_type"] = "dates"
            course["sections"][letter] = [
                {"date": s["date"],
                 "day": date.fromisoformat(s["date"]).isoweekday(),
                 "start": s["start"],
                 "end": s.get("end", SLOT_END.get(s["start"], s["start"]))}
                for s in sorted(sessions, key=lambda x: (x["date"], x["start"]))
            ]
            applied.append(f"{course['code']}-{letter}: "
                           f"{ov.get('note', 'set_dates')} ({len(sessions)} sessions)")
            continue

        if op == "credits":
            credits = float(ov["to"]["credits"])
            course["credits"] = credits
            course.update(CREDIT_RULES[credits])
            applied.append(f"{course['code']}: {ov.get('note', 'credits')}")
            continue

        meetings = course["sections"].get(letter)
        if meetings is None:
            raise SystemExit(f"unknown section {course['code']}-{letter}")

        def find(target, required=True):
            for m in meetings:
                if m.get("day") == target.get("day") and m["start"] == target["start"]:
                    return m
            if not required:
                return None
            raise SystemExit(
                f"override for {course['code']}-{letter} expects a meeting on day "
                f"{target.get('day')} at {target['start']}, which no longer exists. "
                "The published grid may have changed — recheck it."
            )

        if op == "move":
            dest = {"day": ov["to"].get("day", ov["from"]["day"]), "start": ov["to"]["start"]}
            if find(ov["from"], required=False) is None and find(dest, required=False):
                applied.append(f"{course['code']}-{letter}: already applied")
                continue
            m = find(ov["from"])
            m["day"], m["start"] = dest["day"], dest["start"]
            m["end"] = SLOT_END.get(m["start"], m["start"])
        elif op == "remove":
            gone = find(ov["from"], required=False)
            if gone is None:
                applied.append(f"{course['code']}-{letter}: already applied")
                continue
            meetings.remove(gone)
        elif op == "add":
            start = ov["to"]["start"]
            slot = {"day": ov["to"]["day"], "start": start}
            if find(slot, required=False):
                applied.append(f"{course['code']}-{letter}: already applied")
                continue
            meetings.append({**slot, "end": SLOT_END.get(start, start)})
        else:
            raise SystemExit(f"unknown override op '{op}'")

        meetings.sort(key=lambda m: (m.get("date", ""), m.get("day", 0), m["start"]))
        applied.append(f"{course['code']}-{letter}: {ov.get('note', op)}")

    return applied


# ---------------------------------------------------------------- main

if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "data/Class_Schedule_Term-V_AY-2026-27.xlsx"
    dest = sys.argv[2] if len(sys.argv) > 2 else "src/data/catalogue.json"
    overrides = sys.argv[3] if len(sys.argv) > 3 else "data/overrides.json"

    catalogue, matched = build(src)
    applied = apply_overrides(catalogue, overrides)

    with open(dest, "w", encoding="utf-8") as fh:
        json.dump(catalogue, fh, indent=2, ensure_ascii=False)

    weekly = [c for c in catalogue["courses"] if c["schedule_type"] == "weekly"]
    dated = [c for c in catalogue["courses"] if c["schedule_type"] == "dates"]
    meetings = sum(len(v) for c in catalogue["courses"] for v in c["sections"].values())
    print(f"{len(catalogue['courses'])} courses "
          f"({len(weekly)} weekly, {len(dated)} fixed-date), "
          f"{meetings} meetings -> {dest}")

    cal = catalogue["calendar"]
    print(f"\nterm {cal['term_start']} .. {cal['term_end']}  "
          f"(pre-mid ends {cal['pre_mid_end']}, post-mid starts {cal['post_mid_start']})")
    for b in cal["breaks"]:
        print(f"  no class  {b['from']} .. {b['to']}  {b['label']}")

    if matched:
        print("\nfixed-date courses:")
        for code, official, n in matched:
            print(f"  {code:<8} {official:<12} {n} sessions")

    if applied:
        print("\noverrides:")
        for a in applied:
            print("  ", a)

    # A session count that doesn't match the credit rule is the likeliest
    # source of a wrong attendance budget, so say so loudly.
    for c in catalogue["courses"]:
        if c["schedule_type"] != "dates":
            continue
        for letter, ms in c["sections"].items():
            if len(ms) != c["total_classes"]:
                print(f"\n  ! {c['code']}-{letter}: {len(ms)} sessions vs "
                      f"total_classes {c['total_classes']}")