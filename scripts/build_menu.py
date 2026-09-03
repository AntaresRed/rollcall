#!/usr/bin/env python3
"""
Build src/data/menu.json — the day mess menu, one sheet per hostel.

    python3 scripts/build_menu.py \
        "data/Day Mess Menu.xlsx" src/data/menu.json

The workbook is one sheet per hostel, each a week down the side and meals
across the top. Sheet names carry the hostel in front of some variation of
"Day Mess Menu", so the abbreviation is read off the front rather than
hardcoded — a fifth hostel should be a new sheet, not a code change.

Two things the sheets do that a naive reader would lose:

  * OH ends with an "Everyday offering---->" row: bread, eggs, salad, rice and
    so on, served every day and therefore listed against no weekday. Dropping
    it would understate every OH meal. It is split back into per-meal notes
    where the text allows, and kept whole where it does not.
  * Hostels can share a menu — LVH and WH are identical today. They still get
    their own entry, because they are separate messes that happen to agree
    this term, and merging them would need undoing the moment one changes.
"""

import json
import re
import sys

import openpyxl

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday",
        "Friday", "Saturday", "Sunday"]

# The columns as the sheets label them, in the order a day is eaten.
MEALS = ["Breakfast", "Lunch", "Snacks", "Dinner"]

EVERYDAY = re.compile(r"every\s*day\s*offering", re.I)


def norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def die(msg):
    raise SystemExit(f"menu: {msg}")


def hostel_of(sheet_name):
    """'LVH Day Mess Menu' -> 'LVH'."""
    first = norm(sheet_name).split()
    if not first:
        die(f"unnamed sheet")
    return first[0].upper()


def split_everyday(text):
    """'Breakfast: a, b Lunch: c' -> {'Breakfast': 'a, b', 'Lunch': 'c'}.

    Returns the raw text under a single key when the meal labels are not
    there, so an unrecognised note is still shown rather than swallowed.
    """
    body = re.split(r"-{2,}>|:", text, maxsplit=1)
    body = body[-1] if len(body) > 1 else text
    hits = list(re.finditer(r"\b(" + "|".join(MEALS) + r")\s*:", body, re.I))
    if not hits:
        return {"_all": norm(body)}
    out = {}
    for i, m in enumerate(hits):
        end = hits[i + 1].start() if i + 1 < len(hits) else len(body)
        meal = m.group(1).title()
        out[meal] = norm(body[m.end():end]).strip(" ,;")
    return out


def read_sheet(ws):
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(r)]
    if not rows:
        die(f"{ws.title}: empty")

    header = rows[0]
    # Which column each meal sits in, read rather than assumed — a sheet that
    # reorders its columns should still parse, and one that renames them
    # should fail loudly instead of silently emitting blanks.
    cols = {}
    for i, cell in enumerate(header):
        for meal in MEALS:
            if cell.lower() == meal.lower():
                cols[meal] = i
    missing = [m for m in MEALS if m not in cols]
    if missing:
        die(f"{ws.title}: no column for {', '.join(missing)} "
            f"(header reads {header})")

    days, everyday = [], {}
    for r in rows[1:]:
        label = r[0]
        if EVERYDAY.search(" ".join(r)):
            everyday = split_everyday(" ".join(x for x in r if x))
            continue
        match = next((d for d in DAYS if d.lower() == label.lower()), None)
        if not match:
            continue        # a stray note or a blank spacer row
        days.append({
            "day": match,
            "meals": {m: r[cols[m]] if cols[m] < len(r) else "" for m in MEALS},
        })

    got = [d["day"] for d in days]
    if got != DAYS:
        die(f"{ws.title}: expected {DAYS}, found {got}")

    empty = [f"{d['day']} {m}" for d in days for m in MEALS if not d["meals"][m]]
    return days, everyday, empty


def build(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    hostels, warnings = [], []
    for name in wb.sheetnames:
        days, everyday, empty = read_sheet(wb[name])
        hostels.append({
            "id": hostel_of(name).lower(),
            "name": hostel_of(name),
            "everyday": everyday,
            "days": days,
        })
        for e in empty:
            warnings.append(f"{hostel_of(name)}: {e} is blank")

    ids = [h["id"] for h in hostels]
    if len(set(ids)) != len(ids):
        die(f"two sheets name the same hostel: {ids}")

    return {"meals": MEALS, "hostels": hostels}, warnings


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "data/Day Mess Menu.xlsx"
    dest = sys.argv[2] if len(sys.argv) > 2 else "src/data/menu.json"

    menu, warnings = build(src)
    with open(dest, "w", encoding="utf-8", newline="\n") as fh:
        json.dump(menu, fh, indent=2, ensure_ascii=False)

    print(f"{len(menu['hostels'])} hostels -> {dest}")
    for h in menu["hostels"]:
        extra = ""
        if h["everyday"]:
            n = len(h["everyday"])
            extra = f", plus an everyday note ({n} part{'' if n == 1 else 's'})"
        print(f"   {h['name']:<5} {len(h['days'])} days{extra}")

    # Hostels sharing a menu is legitimate but worth saying out loud, so a
    # copy-pasted sheet is noticed rather than shipped as fact.
    seen = {}
    for h in menu["hostels"]:
        key = json.dumps(h["days"], sort_keys=True)
        seen.setdefault(key, []).append(h["name"])
    for names in seen.values():
        if len(names) > 1:
            print(f"\nidentical menus: {' and '.join(names)} — "
                  f"kept as separate tabs, since they are separate messes")

    if warnings:
        print(f"\n{len(warnings)} blank cell(s):")
        for w in warnings:
            print(f"   {w}")
